#!/usr/bin/env python3
"""Rebuild Credit Note template with all improvements from Ant's review."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# === STYLING ===
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
THICK_BORDER_BOTTOM = Border(bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(bold=True, size=20, color='1E40AF')
SUBTITLE_FONT = Font(italic=True, size=9, color='4B5563')
LABEL_FONT = Font(bold=True, size=10, color='374151')
INPUT_FILL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
TOTAL_LABEL_FONT = Font(bold=True, size=11, color='DC2626')
TOTAL_AMOUNT_FONT = Font(bold=True, size=12, color='DC2626')
NOTE_FONT = Font(italic=True, size=9, color='6B7280')
MONEY_FMT = '£#,##0.00;-£#,##0.00;'
PCT_FMT = '0%'

wb = openpyxl.Workbook()

# ============================================================
# COVER SHEET
# ============================================================
ws_cover = wb.active
ws_cover.title = 'Cover'
ws_cover.sheet_view.showGridLines = False

ws_cover.column_dimensions['A'].width = 3
ws_cover.column_dimensions['B'].width = 18
ws_cover.column_dimensions['C'].width = 65

ws_cover['B2'] = 'CREDIT NOTE TEMPLATE'
ws_cover['B2'].font = Font(bold=True, size=18, color='1E40AF')
ws_cover['B3'] = 'Issue credit notes for refunds, returns, and invoice corrections. Maintains accurate VAT records.'
ws_cover['B3'].font = SUBTITLE_FONT

ws_cover['B5'] = 'How to Use'
ws_cover['B5'].font = Font(bold=True, size=12, color='1E40AF')

steps = [
    'Enter your business details in the header section',
    'Add the customer details and original invoice reference',
    'Select a Reason and Credit Applied As from the dropdowns',
    'For each line item, enter description, quantity, rate, and VAT rate (per line)',
    'Net, VAT, and Total calculate automatically per line and in total',
    'Save as PDF and send to your customer — keep a copy for your VAT records',
]
for i, step in enumerate(steps, start=6):
    ws_cover[f'B{i}'] = f'Step {i-5}:'
    ws_cover[f'B{i}'].font = LABEL_FONT
    ws_cover[f'C{i}'] = step
    ws_cover[f'C{i}'].font = Font(size=10)

ws_cover['B13'] = 'Notes'
ws_cover['B13'].font = Font(bold=True, size=12, color='1E40AF')
notes = [
    'Blue cells are for your input — everything else calculates automatically',
    'VAT rates are per line: choose 20% (standard), 5% (reduced), or 0% (zero-rated)',
    'This template is free to use for your business',
    'For support, visit opensheets.pages.dev',
]
for i, note in enumerate(notes, start=14):
    ws_cover[f'B{i}'] = f'• {note}'
    ws_cover[f'B{i}'].font = NOTE_FONT

# ============================================================
# CREDIT NOTE SHEET
# ============================================================
ws = wb.create_sheet('Credit Note')
ws.sheet_view.showGridLines = False

# Column widths — generous to prevent overflow
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 30   # Description / labels
ws.column_dimensions['C'].width = 9    # Qty
ws.column_dimensions['D'].width = 13   # Rate
ws.column_dimensions['E'].width = 11   # VAT Rate
ws.column_dimensions['F'].width = 14   # Net
ws.column_dimensions['G'].width = 14   # VAT
ws.column_dimensions['H'].width = 14   # Total
ws.column_dimensions['I'].width = 3

# --- Title ---
ws['B2'] = 'CREDIT NOTE'
ws['B2'].font = TITLE_FONT
ws['B3'] = 'This document reduces the amount owed by the customer. Keep for VAT records.'
ws['B3'].font = SUBTITLE_FONT

# --- From block ---
ws['B5'] = 'From:'
ws['B5'].font = LABEL_FONT
from_fields = [
    (6, 'YOUR BUSINESS NAME'),
    (7, 'Address Line 1'),
    (8, 'Town, Postcode'),
    (9, 'Tel:'),
    (10, 'Email:'),
    (11, 'VAT Reg No:'),
]
for row, val in from_fields:
    ws[f'B{row}'] = val
    ws[f'B{row}'].fill = INPUT_FILL

# --- Credit Note Details block ---
ws['D5'] = 'Credit Note No:'
ws['D5'].font = LABEL_FONT
ws['E5'] = 'CN-2026-001'
ws['E5'].fill = INPUT_FILL
ws.merge_cells('E5:H5')

ws['D6'] = 'Date:'
ws['D6'].font = LABEL_FONT
ws['E6'] = datetime.now().strftime('%d/%m/%Y')
ws['E6'].fill = INPUT_FILL
ws.merge_cells('E6:H6')

ws['D7'] = 'Original Invoice:'
ws['D7'].font = LABEL_FONT
ws['E7'] = ''
ws['E7'].fill = INPUT_FILL
ws.merge_cells('E7:H7')

ws['D8'] = 'Reason:'
ws['D8'].font = LABEL_FONT
ws['E8'] = ''
ws['E8'].fill = INPUT_FILL
ws.merge_cells('E8:H8')

ws['D9'] = 'Credit Applied As:'
ws['D9'].font = LABEL_FONT
ws['E9'] = ''
ws['E9'].fill = INPUT_FILL
ws.merge_cells('E9:H9')

# --- Dropdown validations ---
dv_reason = DataValidation(type="list", formula1='"Return,Refund,Invoice Correction,Discount,Damage/Defective,Other"', allow_blank=True)
dv_reason.add('E8')
ws.add_data_validation(dv_reason)

dv_applied = DataValidation(type="list", formula1='"Refund to Card/Bank,Credit Against Account,Other"', allow_blank=True)
dv_applied.add('E9')
ws.add_data_validation(dv_applied)

# --- To block ---
ws['B13'] = 'To:'
ws['B13'].font = LABEL_FONT
ws['B14'] = 'Customer Name'
ws['B14'].fill = INPUT_FILL
ws.merge_cells('B14:C14')
ws['B15'] = 'Customer Address'
ws['B15'].fill = INPUT_FILL
ws.merge_cells('B15:C15')

# --- Line item headers ---
hdr_row = 18
headers = [
    ('B', 'Description'),
    ('C', 'Qty'),
    ('D', 'Rate (£)'),
    ('E', 'VAT Rate'),
    ('F', 'Net (£)'),
    ('G', 'VAT (£)'),
    ('H', 'Total (£)'),
]
for col, val in headers:
    cell = ws[f'{col}{hdr_row}']
    cell.value = val
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')

# --- Line items (rows 19–24) ---
dv_vat = DataValidation(type="list", formula1='"20%,5%,0%"', allow_blank=True)
for row in range(19, 25):
    dv_vat.add(f'E{row}')
    
    # Description
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].border = THIN_BORDER
    
    # Qty
    ws[f'C{row}'] = ''
    ws[f'C{row}'].fill = INPUT_FILL
    ws[f'C{row}'].border = THIN_BORDER
    ws[f'C{row}'].alignment = Alignment(horizontal='center')
    
    # Rate
    ws[f'D{row}'] = ''
    ws[f'D{row}'].fill = INPUT_FILL
    ws[f'D{row}'].border = THIN_BORDER
    ws[f'D{row}'].number_format = MONEY_FMT
    
    # VAT Rate
    ws[f'E{row}'] = '20%'
    ws[f'E{row}'].fill = INPUT_FILL
    ws[f'E{row}'].border = THIN_BORDER
    ws[f'E{row}'].alignment = Alignment(horizontal='center')
    
    # Net = IF(blank or zero, blank, qty*rate)
    ws[f'F{row}'] = f'=IF(OR(C{row}="",D{row}="",C{row}=0,D{row}=0),"",C{row}*D{row})'
    ws[f'F{row}'].border = THIN_BORDER
    ws[f'F{row}'].number_format = MONEY_FMT
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    
    # VAT = IF(net blank or vat blank, blank, net * parsed_vat_rate)
    vat_formula = (
        f'=IF(OR(F{row}="",E{row}=""),"",'
        f'F{row}*IF(E{row}="20%",0.2,IF(E{row}="5%",0.05,0)))'
    )
    ws[f'G{row}'] = vat_formula
    ws[f'G{row}'].border = THIN_BORDER
    ws[f'G{row}'].number_format = MONEY_FMT
    ws[f'G{row}'].alignment = Alignment(horizontal='right')
    
    # Total = net + vat
    ws[f'H{row}'] = f'=IF(OR(F{row}="",G{row}=""),"",F{row}+G{row})'
    ws[f'H{row}'].border = THIN_BORDER
    ws[f'H{row}'].number_format = MONEY_FMT
    ws[f'H{row}'].alignment = Alignment(horizontal='right')

ws.add_data_validation(dv_vat)

# --- Totals section ---
total_row = 26
ws[f'B{total_row}'] = ''
ws[f'F{total_row}'] = 'NET TOTAL'
ws[f'F{total_row}'].font = LABEL_FONT
ws[f'F{total_row}'].alignment = Alignment(horizontal='right')
ws[f'G{total_row}'] = 'VAT TOTAL'
ws[f'G{total_row}'].font = LABEL_FONT
ws[f'G{total_row}'].alignment = Alignment(horizontal='right')
ws[f'H{total_row}'] = 'TOTAL CREDIT'
ws[f'H{total_row}'].font = TOTAL_LABEL_FONT
ws[f'H{total_row}'].alignment = Alignment(horizontal='right')

ws[f'F{total_row+1}'] = '=SUM(F19:F24)'
ws[f'F{total_row+1}'].font = LABEL_FONT
ws[f'F{total_row+1}'].number_format = MONEY_FMT
ws[f'F{total_row+1}'].alignment = Alignment(horizontal='right')
ws[f'F{total_row+1}'].border = Border(top=Side(style='thin'), bottom=Side(style='double'))

ws[f'G{total_row+1}'] = '=SUM(G19:G24)'
ws[f'G{total_row+1}'].font = LABEL_FONT
ws[f'G{total_row+1}'].number_format = MONEY_FMT
ws[f'G{total_row+1}'].alignment = Alignment(horizontal='right')
ws[f'G{total_row+1}'].border = Border(top=Side(style='thin'), bottom=Side(style='double'))

ws[f'H{total_row+1}'] = '=F27+G27'
ws[f'H{total_row+1}'].font = TOTAL_AMOUNT_FONT
ws[f'H{total_row+1}'].number_format = MONEY_FMT
ws[f'H{total_row+1}'].alignment = Alignment(horizontal='right')
ws[f'H{total_row+1}'].border = Border(top=Side(style='thin'), bottom=Side(style='double'))

# --- Notes / Footer ---
ws['B29'] = 'Notes:'
ws['B29'].font = LABEL_FONT
ws['B30'] = '• This credit note is valid for VAT purposes. Keep a copy for your records.'
ws['B31'] = '• VAT rates: 20% = standard rate, 5% = reduced rate, 0% = zero-rated.'
ws['B32'] = '• For questions, visit opensheets.pages.dev'
for r in range(30, 33):
    ws[f'B{r}'].font = NOTE_FONT

# --- Signature section ---
ws['B34'] = 'Issued By:'
ws['B34'].font = LABEL_FONT
ws['C34'] = ''
ws['C34'].fill = INPUT_FILL
ws['B35'] = 'Date:'
ws['B35'].font = LABEL_FONT
ws['C35'] = ''
ws['C35'].fill = INPUT_FILL
ws['B36'] = 'Authorised Signature:'
ws['B36'].font = LABEL_FONT
ws['C36'] = ''
ws['C36'].fill = INPUT_FILL

# Save
wb.save('app/public/templates/Credit_Note_Template.xlsx')
print('✅ Credit Note Template rebuilt with all improvements!')
print('   • Line-level VAT rates (20% / 5% / 0%) with dropdowns')
print('   • £ currency format with blank zeros')
print('   • Reason & Credit Applied As dropdowns')
print('   • No duplicate reason fields')
print('   • Navy title, red TOTAL CREDIT')
print('   • Gridlines hidden, clean borders')
