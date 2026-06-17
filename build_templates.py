#!/usr/bin/env python3
"""Build OpenSheets templates rapidly."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# Common styling
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=18, color='1E40AF')
SUBTITLE_FONT = Font(bold=True, size=12, color='1E40AF')
LABEL_FONT = Font(bold=True, size=10)
MONEY_FONT = Font(size=10)
BLUE_INPUT = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
LIGHT_GREY = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
NOTE_FONT = Font(italic=True, size=9, color='6B7280')

def add_cover(ws, title, description, steps):
    """Add a standard cover sheet."""
    ws['B2'] = title
    ws['B2'].font = TITLE_FONT
    ws['B3'] = description
    ws['B3'].font = Font(size=11, color='4B5563')
    
    ws['B5'] = 'How to Use'
    ws['B5'].font = SUBTITLE_FONT
    
    for i, step in enumerate(steps, start=6):
        ws[f'B{i}'] = f"Step {i-5}:"
        ws[f'B{i}'].font = LABEL_FONT
        ws[f'C{i}'] = step
        ws[f'C{i}'].font = Font(size=10)
    
    ws['B13'] = 'Notes'
    ws['B13'].font = SUBTITLE_FONT
    ws['B14'] = '• Blue cells are for your input — everything else calculates automatically'
    ws['B15'] = '• This template is free to use for your business'
    ws['B16'] = '• For support, visit opensheets.pages.dev'
    for r in range(14, 17):
        ws[f'B{r}'].font = NOTE_FONT
    
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 60

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_range(ws, cell_range, font=None, fill=None, border=None, alignment=None):
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if alignment: cell.alignment = alignment

# ============================================================
# 1. PRO-FORMA INVOICE
# ============================================================
wb1 = openpyxl.Workbook()
ws_cover = wb1.active
ws_cover.title = 'Cover'
add_cover(ws_cover, 'PRO-FORMA INVOICE',
    'A pro-forma invoice is sent BEFORE goods or services are delivered. It is not a demand for payment but a quotation or pre-payment request.',
    [
        'Enter your business details and logo in the header section',
        'Add customer details and your pro-forma reference number',
        'List items, quantities, and rates — totals calculate automatically',
        'Add payment terms and bank details if requesting pre-payment',
        'Save as PDF and send to your customer'
    ])

ws = wb1.create_sheet('Pro-Forma Invoice')
set_col_widths(ws, {'A':3, 'B':35, 'C':12, 'D':15, 'E':18, 'F':3})

ws['B2'] = 'PRO-FORMA INVOICE'
ws['B2'].font = Font(bold=True, size=20, color='1E40AF')
ws['B3'] = 'This is NOT a tax invoice. Payment terms and details are for guidance only.'
ws['B3'].font = Font(italic=True, size=9, color='DC2626')

# Business details
ws['B5'] = 'From:'
ws['B5'].font = LABEL_FONT
ws['B6'] = 'YOUR BUSINESS NAME'
ws['B6'].fill = BLUE_INPUT
ws['B7'] = 'Address Line 1'
ws['B7'].fill = BLUE_INPUT
ws['B8'] = 'Town, Postcode'
ws['B8'].fill = BLUE_INPUT
ws['B9'] = 'Tel:'
ws['B9'].fill = BLUE_INPUT
ws['B10'] = 'Email:'
ws['B10'].fill = BLUE_INPUT
ws['B11'] = 'VAT Reg No:'
ws['B11'].fill = BLUE_INPUT

# Invoice details
ws['D5'] = 'Pro-Forma Ref:'
ws['D5'].font = LABEL_FONT
ws['E5'] = 'PF-2026-001'
ws['E5'].fill = BLUE_INPUT
ws['D6'] = 'Date:'
ws['D6'].font = LABEL_FONT
ws['E6'] = datetime.now().strftime('%d/%m/%Y')
ws['E6'].fill = BLUE_INPUT
ws['D7'] = 'Valid Until:'
ws['D7'].font = LABEL_FONT
ws['E7'] = ''
ws['E7'].fill = BLUE_INPUT
ws['D8'] = 'Project:'
ws['D8'].font = LABEL_FONT
ws['E8'] = ''
ws['E8'].fill = BLUE_INPUT

# Customer
ws['B13'] = 'TO:'
ws['B13'].font = LABEL_FONT
ws['B14'] = 'Customer Name'
ws['B14'].fill = BLUE_INPUT
ws['B15'] = 'Customer Address'
ws['B15'].fill = BLUE_INPUT

# Line items header
for col, val in [('B', 'Description'), ('C', 'Qty'), ('D', 'Rate (£)'), ('E', 'Amount (£)')]:
    ws[f'{col}18'] = val
    ws[f'{col}18'].font = HEADER_FONT
    ws[f'{col}18'].fill = HEADER_FILL
    ws[f'{col}18'].border = THIN_BORDER
    ws[f'{col}18'].alignment = Alignment(horizontal='center')

# Line items (rows 19-24)
for row in range(19, 25):
    ws[f'B{row}'].fill = BLUE_INPUT
    ws[f'B{row}'].border = THIN_BORDER
    ws[f'C{row}'] = 0
    ws[f'C{row}'].fill = BLUE_INPUT
    ws[f'C{row}'].border = THIN_BORDER
    ws[f'C{row}'].alignment = Alignment(horizontal='center')
    ws[f'D{row}'] = 0
    ws[f'D{row}'].fill = BLUE_INPUT
    ws[f'D{row}'].border = THIN_BORDER
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'E{row}'] = f'=C{row}*D{row}'
    ws[f'E{row}'].border = THIN_BORDER
    ws[f'E{row}'].number_format = '#,##0.00'

# Totals
ws['D26'] = 'SUBTOTAL'
ws['D26'].font = LABEL_FONT
ws['E26'] = '=SUM(E19:E24)'
ws['E26'].font = LABEL_FONT
ws['E26'].number_format = '#,##0.00'
ws['D27'] = 'VAT @ 20%'
ws['D27'].font = LABEL_FONT
ws['E27'] = '=E26*0.2'
ws['E27'].font = LABEL_FONT
ws['E27'].number_format = '#,##0.00'
ws['D28'] = 'TOTAL'
ws['D28'].font = Font(bold=True, size=12)
ws['E28'] = '=E26+E27'
ws['E28'].font = Font(bold=True, size=12)
ws['E28'].number_format = '#,##0.00'

# Payment details
ws['B30'] = 'Payment Terms:'
ws['B30'].font = LABEL_FONT
ws['B31'] = 'Bank:'
ws['B31'].fill = BLUE_INPUT
ws['B32'] = 'Sort Code:'
ws['B32'].fill = BLUE_INPUT
ws['B33'] = 'Account:'
ws['B33'].fill = BLUE_INPUT

ws['B35'] = 'Notes:'
ws['B35'].font = LABEL_FONT
ws['B36'] = 'This pro-forma is valid for 30 days. A formal tax invoice will be issued upon payment or delivery.'
ws['B36'].font = NOTE_FONT

wb1.save('app/public/templates/Pro_Forma_Invoice.xlsx')
print('✅ Pro-Forma Invoice built')

# ============================================================
# 2. QUOTE / ESTIMATE TEMPLATE
# ============================================================
wb2 = openpyxl.Workbook()
ws_cover = wb2.active
ws_cover.title = 'Cover'
add_cover(ws_cover, 'QUOTE / ESTIMATE TEMPLATE',
    'Send professional quotes and estimates to customers. Automatically calculates totals, VAT, and markup. Converts to an invoice when accepted.',
    [
        'Enter your business details in the header',
        'Add customer details and quote reference',
        'List items with optional markup percentage',
        'Set VAT rate (default 20%, change if needed)',
        'Add terms and expiry date, then save as PDF'
    ])

ws = wb2.create_sheet('Quote')
set_col_widths(ws, {'A':3, 'B':30, 'C':10, 'D':14, 'E':14, 'F':16, 'G':3})

ws['B2'] = 'QUOTE'
ws['B2'].font = Font(bold=True, size=20, color='1E40AF')
ws['C2'] = '/ ESTIMATE'
ws['C2'].font = Font(bold=True, size=14, color='6B7280')

ws['B4'] = 'From:'
ws['B4'].font = LABEL_FONT
for r, label in [(5,'YOUR BUSINESS NAME'), (6,'Address'), (7,'Town, Postcode'), (8,'Tel:'), (9,'Email:')]:
    ws[f'B{r}'] = label
    ws[f'B{r}'].fill = BLUE_INPUT

ws['D4'] = 'Quote Ref:'
ws['D4'].font = LABEL_FONT
ws['E4'] = 'Q-2026-001'
ws['E4'].fill = BLUE_INPUT
ws['D5'] = 'Date:'
ws['D5'].font = LABEL_FONT
ws['E5'] = datetime.now().strftime('%d/%m/%Y')
ws['E5'].fill = BLUE_INPUT
ws['D6'] = 'Valid Until:'
ws['D6'].font = LABEL_FONT
ws['E6'] = ''
ws['E6'].fill = BLUE_INPUT
ws['D7'] = 'Customer:'
ws['D7'].font = LABEL_FONT
ws['E7'] = ''
ws['E7'].fill = BLUE_INPUT

# Headers
headers = [('B','Description'), ('C','Qty'), ('D','Cost (£)'), ('E','Markup %'), ('F','Total (£)')]
for col, val in headers:
    ws[f'{col}11'] = val
    ws[f'{col}11'].font = HEADER_FONT
    ws[f'{col}11'].fill = HEADER_FILL
    ws[f'{col}11'].border = THIN_BORDER
    ws[f'{col}11'].alignment = Alignment(horizontal='center')

# Line items (12-17)
for row in range(12, 18):
    ws[f'B{row}'].fill = BLUE_INPUT
    ws[f'B{row}'].border = THIN_BORDER
    ws[f'C{row}'] = 1
    ws[f'C{row}'].fill = BLUE_INPUT
    ws[f'C{row}'].border = THIN_BORDER
    ws[f'D{row}'] = 0
    ws[f'D{row}'].fill = BLUE_INPUT
    ws[f'D{row}'].border = THIN_BORDER
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'E{row}'] = 0
    ws[f'E{row}'].fill = BLUE_INPUT
    ws[f'E{row}'].border = THIN_BORDER
    ws[f'F{row}'] = f'=C{row}*D{row}*(1+E{row})'
    ws[f'F{row}'].border = THIN_BORDER
    ws[f'F{row}'].number_format = '#,##0.00'

ws['E19'] = 'SUBTOTAL'
ws['E19'].font = LABEL_FONT
ws['F19'] = '=SUM(F12:F17)'
ws['F19'].font = LABEL_FONT
ws['F19'].number_format = '#,##0.00'

ws['E20'] = 'VAT Rate:'
ws['E20'].font = LABEL_FONT
ws['F20'] = 0.20
ws['F20'].fill = BLUE_INPUT
ws['F20'].number_format = '0%'

ws['E21'] = 'VAT Amount'
ws['E21'].font = LABEL_FONT
ws['F21'] = '=F19*F20'
ws['F21'].number_format = '#,##0.00'

ws['E22'] = 'QUOTE TOTAL'
ws['E22'].font = Font(bold=True, size=12)
ws['F22'] = '=F19+F21'
ws['F22'].font = Font(bold=True, size=12)
ws['F22'].number_format = '#,##0.00'

ws['B24'] = 'Terms & Conditions:'
ws['B24'].font = LABEL_FONT
ws['B25'] = '• This quote is valid for 30 days from the date of issue'
ws['B26'] = '• Payment terms: 50% deposit, 50% on completion'
ws['B27'] = '• All prices are in GBP (£)'
ws['B28'] = '• Work will commence upon written acceptance of this quote'
for r in range(25, 29):
    ws[f'B{r}'].font = NOTE_FONT

wb2.save('app/public/templates/Quote_Estimate_Template.xlsx')
print('✅ Quote/Estimate built')

# ============================================================
# 3. CREDIT NOTE TEMPLATE
# ============================================================
wb3 = openpyxl.Workbook()
ws_cover = wb3.active
ws_cover.title = 'Cover'
add_cover(ws_cover, 'CREDIT NOTE TEMPLATE',
    'Issue credit notes for returns, refunds, or invoice corrections. Essential for maintaining accurate VAT records.',
    [
        'Enter your business details and the original invoice reference',
        'State the reason for the credit note',
        'List items being credited with original and corrected values',
        'VAT is calculated automatically on the credit amount',
        'Keep a copy for your VAT records'
    ])

ws = wb3.create_sheet('Credit Note')
set_col_widths(ws, {'A':3, 'B':35, 'C':12, 'D':15, 'E':18, 'F':3})

ws['B2'] = 'CREDIT NOTE'
ws['B2'].font = Font(bold=True, size=20, color='DC2626')
ws['B3'] = 'This document reduces the amount owed by the customer. Keep for VAT records.'
ws['B3'].font = Font(italic=True, size=9, color='DC2626')

ws['B5'] = 'From:'
ws['B5'].font = LABEL_FONT
for r, label in [(6,'YOUR BUSINESS NAME'), (7,'Address'), (8,'Town, Postcode'), (9,'Tel:'), (10,'Email:'), (11,'VAT Reg No:')]:
    ws[f'B{r}'] = label
    ws[f'B{r}'].fill = BLUE_INPUT

ws['D5'] = 'Credit Note No:'
ws['D5'].font = LABEL_FONT
ws['E5'] = 'CN-2026-001'
ws['E5'].fill = BLUE_INPUT
ws['D6'] = 'Date:'
ws['D6'].font = LABEL_FONT
ws['E6'] = datetime.now().strftime('%d/%m/%Y')
ws['E6'].fill = BLUE_INPUT
ws['D7'] = 'Original Invoice:'
ws['D7'].font = LABEL_FONT
ws['E7'] = ''
ws['E7'].fill = BLUE_INPUT
ws['D8'] = 'Reason:'
ws['D8'].font = LABEL_FONT
ws['E8'] = 'Return / Refund / Correction'
ws['E8'].fill = BLUE_INPUT

ws['B12'] = 'TO:'
ws['B12'].font = LABEL_FONT
ws['B13'] = 'Customer Name'
ws['B13'].fill = BLUE_INPUT
ws['B14'] = 'Customer Address'
ws['B14'].fill = BLUE_INPUT

# Headers
for col, val in [('B', 'Description'), ('C', 'Qty'), ('D', 'Rate (£)'), ('E', 'Credit Amount (£)')]:
    ws[f'{col}17'] = val
    ws[f'{col}17'].font = HEADER_FONT
    ws[f'{col}17'].fill = HEADER_FILL
    ws[f'{col}17'].border = THIN_BORDER
    ws[f'{col}17'].alignment = Alignment(horizontal='center')

for row in range(18, 23):
    ws[f'B{row}'].fill = BLUE_INPUT
    ws[f'B{row}'].border = THIN_BORDER
    ws[f'C{row}'] = 0
    ws[f'C{row}'].fill = BLUE_INPUT
    ws[f'C{row}'].border = THIN_BORDER
    ws[f'D{row}'] = 0
    ws[f'D{row}'].fill = BLUE_INPUT
    ws[f'D{row}'].border = THIN_BORDER
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'E{row}'] = f'=C{row}*D{row}'
    ws[f'E{row}'].border = THIN_BORDER
    ws[f'E{row}'].number_format = '#,##0.00'

ws['D24'] = 'SUBTOTAL'
ws['D24'].font = LABEL_FONT
ws['E24'] = '=SUM(E18:E22)'
ws['E24'].font = LABEL_FONT
ws['E24'].number_format = '#,##0.00'

ws['D25'] = 'VAT @ 20%'
ws['D25'].font = LABEL_FONT
ws['E25'] = '=E24*0.2'
ws['E25'].font = LABEL_FONT
ws['E25'].number_format = '#,##0.00'

ws['D26'] = 'TOTAL CREDIT'
ws['D26'].font = Font(bold=True, size=12, color='DC2626')
ws['E26'] = '=E24+E25'
ws['E26'].font = Font(bold=True, size=12, color='DC2626')
ws['E26'].number_format = '#,##0.00'

ws['B28'] = 'Reason for Credit:'
ws['B28'].font = LABEL_FONT
ws['B29'] = ''
ws['B29'].fill = BLUE_INPUT

wb3.save('app/public/templates/Credit_Note_Template.xlsx')
print('✅ Credit Note built')

# ============================================================
# 4. DELIVERY NOTE TEMPLATE
# ============================================================
wb4 = openpyxl.Workbook()
ws_cover = wb4.active
ws_cover.title = 'Cover'
add_cover(ws_cover, 'DELIVERY NOTE TEMPLATE',
    'Accompany goods sent to customers with a professional delivery note. Lists items, quantities, and condition. No prices shown.',
    [
        'Enter your business details and delivery reference',
        'Add customer and delivery address',
        'List items and quantities delivered',
        'Record who received the goods and the date',
        'Get a signature — this is proof of delivery'
    ])

ws = wb4.create_sheet('Delivery Note')
set_col_widths(ws, {'A':3, 'B':40, 'C':12, 'D':20, 'E':18, 'F':3})

ws['B2'] = 'DELIVERY NOTE'
ws['B2'].font = Font(bold=True, size=20, color='1E40AF')

ws['B4'] = 'From:'
ws['B4'].font = LABEL_FONT
for r, label in [(5,'YOUR BUSINESS NAME'), (6,'Address'), (7,'Town, Postcode'), (8,'Tel:')]:
    ws[f'B{r}'] = label
    ws[f'B{r}'].fill = BLUE_INPUT

ws['D4'] = 'Delivery Note No:'
ws['D4'].font = LABEL_FONT
ws['E4'] = 'DN-2026-001'
ws['E4'].fill = BLUE_INPUT
ws['D5'] = 'Date:'
ws['D5'].font = LABEL_FONT
ws['E5'] = datetime.now().strftime('%d/%m/%Y')
ws['E5'].fill = BLUE_INPUT
ws['D6'] = 'Order Ref:'
ws['D6'].font = LABEL_FONT
ws['E6'] = ''
ws['E6'].fill = BLUE_INPUT
ws['D7'] = 'Customer:'
ws['D7'].font = LABEL_FONT
ws['E7'] = ''
ws['E7'].fill = BLUE_INPUT

ws['B9'] = 'Deliver To:'
ws['B9'].font = LABEL_FONT
ws['B10'] = 'Delivery Address'
ws['B10'].fill = BLUE_INPUT
ws['B11'] = ''
ws['B11'].fill = BLUE_INPUT

# Headers
for col, val in [('B', 'Item Description'), ('C', 'Qty'), ('D', 'Condition'), ('E', 'Notes')]:
    ws[f'{col}13'] = val
    ws[f'{col}13'].font = HEADER_FONT
    ws[f'{col}13'].fill = HEADER_FILL
    ws[f'{col}13'].border = THIN_BORDER
    ws[f'{col}13'].alignment = Alignment(horizontal='center')

for row in range(14, 20):
    ws[f'B{row}'].fill = BLUE_INPUT
    ws[f'B{row}'].border = THIN_BORDER
    ws[f'C{row}'] = 0
    ws[f'C{row}'].fill = BLUE_INPUT
    ws[f'C{row}'].border = THIN_BORDER
    ws[f'D{row}'] = 'Good / Damaged'
    ws[f'D{row}'].fill = BLUE_INPUT
    ws[f'D{row}'].border = THIN_BORDER
    ws[f'E{row}'] = ''
    ws[f'E{row}'].fill = BLUE_INPUT
    ws[f'E{row}'].border = THIN_BORDER

ws['B22'] = 'Total Items:'
ws['B22'].font = LABEL_FONT
ws['C22'] = '=SUM(C14:C19)'
ws['C22'].font = LABEL_FONT

ws['B24'] = 'Received By:'
ws['B24'].font = LABEL_FONT
ws['C24'] = ''
ws['C24'].fill = BLUE_INPUT
ws['B25'] = 'Signature:'
ws['B25'].font = LABEL_FONT
ws['C25'] = ''
ws['C25'].fill = BLUE_INPUT
ws['B26'] = 'Date Received:'
ws['B26'].font = LABEL_FONT
ws['C26'] = ''
ws['C26'].fill = BLUE_INPUT

wb4.save('app/public/templates/Delivery_Note_Template.xlsx')
print('✅ Delivery Note built')

# ============================================================
# 5. WEEKLY TIMESHEET
# ============================================================
wb5 = openpyxl.Workbook()
ws_cover = wb5.active
ws_cover.title = 'Cover'
add_cover(ws_cover, 'WEEKLY TIMESHEET',
    'Track employee or contractor hours by day. Calculates regular hours, overtime, and total pay automatically.',
    [
        'Enter employee name and pay rate in the header',
        'Record start and finish times for each day (use 24h format)',
        'Enter break duration in minutes',
        'Overtime kicks in after 40 hours per week (customisable)',
        'Total pay calculates automatically'
    ])

ws = wb5.create_sheet('Timesheet')
set_col_widths(ws, {'A':3, 'B':14, 'C':14, 'D':14, 'E':14, 'F':12, 'G':14, 'H':14, 'I':3})

ws['B2'] = 'WEEKLY TIMESHEET'
ws['B2'].font = Font(bold=True, size=20, color='1E40AF')

ws['B4'] = 'Employee Name:'
ws['B4'].font = LABEL_FONT
ws['C4'] = ''
ws['C4'].fill = BLUE_INPUT
ws['B5'] = 'Week Ending:'
ws['B5'].font = LABEL_FONT
ws['C5'] = ''
ws['C5'].fill = BLUE_INPUT
ws['B6'] = 'Hourly Rate (£):'
ws['B6'].font = LABEL_FONT
ws['C6'] = 0
ws['C6'].fill = BLUE_INPUT
ws['C6'].number_format = '#,##0.00'
ws['B7'] = 'Overtime Rate (x):'
ws['B7'].font = LABEL_FONT
ws['C7'] = 1.5
ws['C7'].fill = BLUE_INPUT
ws['C7'].number_format = '0.00'

# Headers
headers = [('B','Day'), ('C','Start'), ('D','Finish'), ('E','Break (min)'), ('F','Hours'), ('G','Notes')]
for col, val in headers:
    ws[f'{col}9'] = val
    ws[f'{col}9'].font = HEADER_FONT
    ws[f'{col}9'].fill = HEADER_FILL
    ws[f'{col}9'].border = THIN_BORDER
    ws[f'{col}9'].alignment = Alignment(horizontal='center')

days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
for i, day in enumerate(days):
    row = 10 + i
    ws[f'B{row}'] = day
    ws[f'B{row}'].border = THIN_BORDER
    ws[f'C{row}'] = ''
    ws[f'C{row}'].fill = BLUE_INPUT
    ws[f'C{row}'].border = THIN_BORDER
    ws[f'D{row}'] = ''
    ws[f'D{row}'].fill = BLUE_INPUT
    ws[f'D{row}'].border = THIN_BORDER
    ws[f'E{row}'] = 0
    ws[f'E{row}'].fill = BLUE_INPUT
    ws[f'E{row}'].border = THIN_BORDER
    ws[f'E{row}'].alignment = Alignment(horizontal='center')
    # Hours = (Finish - Start) - Break/60
    ws[f'F{row}'] = f'=IF(OR(C{row}="",D{row}=""),0,MAX(0,(D{row}-C{row})*24-E{row}/60))'
    ws[f'F{row}'].border = THIN_BORDER
    ws[f'F{row}'].number_format = '0.00'
    ws[f'G{row}'] = ''
    ws[f'G{row}'].fill = BLUE_INPUT
    ws[f'G{row}'].border = THIN_BORDER

# Totals
ws['B18'] = 'TOTALS'
ws['B18'].font = LABEL_FONT
ws['F18'] = '=SUM(F10:F16)'
ws['F18'].font = LABEL_FONT
ws['F18'].number_format = '0.00'

ws['B20'] = 'Regular Hours (up to 40):'
ws['B20'].font = LABEL_FONT
ws['F20'] = '=MIN(F18,40)'
ws['F20'].number_format = '0.00'

ws['B21'] = 'Overtime Hours:'
ws['B21'].font = LABEL_FONT
ws['F21'] = '=MAX(0,F18-40)'
ws['F21'].number_format = '0.00'

ws['B22'] = 'Regular Pay (£):'
ws['B22'].font = LABEL_FONT
ws['F22'] = '=F20*C6'
ws['F22'].number_format = '#,##0.00'

ws['B23'] = 'Overtime Pay (£):'
ws['B23'].font = LABEL_FONT
ws['F23'] = '=F21*C6*C7'
ws['F23'].number_format = '#,##0.00'

ws['B24'] = 'TOTAL PAY (£):'
ws['B24'].font = Font(bold=True, size=12)
ws['F24'] = '=F22+F23'
ws['F24'].font = Font(bold=True, size=12)
ws['F24'].number_format = '#,##0.00'

ws['B26'] = 'Employee Signature:'
ws['B26'].font = LABEL_FONT
ws['C26'] = ''
ws['C26'].fill = BLUE_INPUT
ws['B27'] = 'Manager Signature:'
ws['B27'].font = LABEL_FONT
ws['C27'] = ''
ws['C27'].fill = BLUE_INPUT

wb5.save('app/public/templates/Weekly_Timesheet.xlsx')
print('✅ Weekly Timesheet built')

print('\n🎉 ALL 5 TEMPLATES BUILT SUCCESSFULLY!')
