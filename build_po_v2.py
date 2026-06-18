#!/usr/bin/env python3
"""
Build Purchase Order template per business-doc-template-v2 skill spec.
Navy house style, A4 print-ready, field-prompt party card, blue/grey table,
per-line VAT dropdowns, Cover sheet with OpenSheets + Aligned.tax promos.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

# ============================================================
# PALETTE
# ============================================================
NAVY = '1E40AF'
CHARCOAL = '1F2937'
INPUT_BLUE = 'DBEAFE'
READONLY_GREY = 'F3F4F6'
BORDER_GREY = '9CA3AF'
SEPARATOR_GREY = 'CBD5E1'
ROW_SEP = 'E5E7EB'
MUTED = '6B7280'
LABEL_GREY = '374151'
PROMO_BG = 'EFF6FF'

# ============================================================
# BORDERS
# ============================================================
def thin_border(c=BORDER_GREY):
    return Border(
        left=Side(style='thin', color=c),
        right=Side(style='thin', color=c),
        top=Side(style='thin', color=c),
        bottom=Side(style='thin', color=c)
    )

def outer_only(c=BORDER_GREY):
    return Border(
        left=Side(style='thin', color=c),
        right=Side(style='thin', color=c),
        top=Side(style='thin', color=c),
        bottom=Side(style='thin', color=c)
    )

def sep_h(c=SEPARATOR_GREY):
    return Border(bottom=Side(style='thin', color=c))

def row_sep(c=ROW_SEP):
    return Border(bottom=Side(style='thin', color=c))

def navy_top():
    return Border(top=Side(style='medium', color=NAVY))

def thin_bottom(c=ROW_SEP):
    return Border(bottom=Side(style='thin', color=c))

def no_border():
    return Border()

# ============================================================
# FONTS
# ============================================================
TITLE_FONT = Font(name='Calibri', bold=True, size=26, color=CHARCOAL)
SUBTITLE_FONT = Font(name='Calibri', size=11, color=MUTED)
BAND_LABEL_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
BAND_VALUE_FONT = Font(name='Calibri', bold=True, size=18, color='FFFFFF')
SECTION_LABEL_FONT = Font(name='Calibri', bold=True, size=11, color=LABEL_GREY)
DETAIL_LABEL_FONT = Font(name='Calibri', bold=True, size=12, color=LABEL_GREY)
DETAIL_VALUE_FONT = Font(name='Calibri', size=13, color=CHARCOAL)
PARTY_PROMPT_FONT = Font(name='Calibri', italic=True, size=12, color=MUTED)
TABLE_HEADER_FONT = Font(name='Calibri', bold=True, size=13, color=CHARCOAL)
TABLE_BODY_FONT = Font(name='Calibri', size=13, color=CHARCOAL)
TOTAL_LABEL_FONT = Font(name='Calibri', bold=True, size=13, color=CHARCOAL)
TOTAL_AMOUNT_FONT = Font(name='Calibri', bold=True, size=15, color=CHARCOAL)
FOOTER_FONT = Font(name='Calibri', size=11, color=CHARCOAL)
NOTE_FONT = Font(name='Calibri', italic=True, size=11, color=MUTED)

# ============================================================
# FILLS
# ============================================================
INPUT_FILL = PatternFill(start_color=INPUT_BLUE, end_color=INPUT_BLUE, fill_type='solid')
READONLY_FILL = PatternFill(start_color=READONLY_GREY, end_color=READONLY_GREY, fill_type='solid')
NAVY_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
CHARCOAL_FILL = PatternFill(start_color=CHARCOAL, end_color=CHARCOAL, fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
PROMO_FILL = PatternFill(start_color=PROMO_BG, end_color=PROMO_BG, fill_type='solid')

# ============================================================
# NUMBER FORMATS
# ============================================================
MONEY_FMT = '£#,##0.00;-£#,##0.00;"-"'
QTY_FMT = '#,##0;-#,##0;"-"'
PCT_FMT = '0%'
DATE_FMT = 'DD/MM/YYYY'

# ============================================================
# ALIGNMENT
# ============================================================
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_TOP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_TOP_RIGHT = Alignment(horizontal='right', vertical='top')

# ============================================================
# UNLOCKED CELL (for protection)
# ============================================================
UNLOCKED = Protection(locked=False)

wb = openpyxl.Workbook()

# ============================================================
# COVER SHEET
# ============================================================
ws_cover = wb.active
ws_cover.title = 'Cover'
ws_cover.sheet_view.showGridLines = False

# Column widths
ws_cover.column_dimensions['A'].width = 3
ws_cover.column_dimensions['B'].width = 20
ws_cover.column_dimensions['C'].width = 55
ws_cover.column_dimensions['D'].width = 3

# Title
ws_cover['B2'] = 'PURCHASE ORDER TEMPLATE'
ws_cover['B2'].font = Font(name='Calibri', bold=True, size=18, color=NAVY)
ws_cover['B3'] = 'Send professional purchase orders to suppliers. Per-line VAT rates, auto-totals, and print-ready layout.'
ws_cover['B3'].font = SUBTITLE_FONT

# How to Use
ws_cover['B5'] = 'How to Use'
ws_cover['B5'].font = Font(name='Calibri', bold=True, size=12, color=NAVY)

steps = [
    'Enter your company details in the footer (single source of truth — auto-populates the top)',
    'Fill in supplier details in the left card and PO details in the right card',
    'Add line items: description, quantity, unit price, and select VAT rate per line',
    'Net, VAT, and Total calculate automatically for each line and in the totals block',
    'Add any notes or delivery instructions, then print or save as PDF',
]
for i, step in enumerate(steps, start=6):
    ws_cover[f'B{i}'] = f'{i-5}.'
    ws_cover[f'B{i}'].font = Font(name='Calibri', bold=True, size=10, color=LABEL_GREY)
    ws_cover[f'C{i}'] = step
    ws_cover[f'C{i}'].font = Font(name='Calibri', size=10, color=CHARCOAL)

# Notes
ws_cover['B12'] = 'Notes'
ws_cover['B12'].font = Font(name='Calibri', bold=True, size=12, color=NAVY)
notes = [
    'Blue cells are for your input — totals calculate automatically',
    'Grey cells are read-only — do not edit them directly',
    'VAT rates are per line: choose 20% (standard), 5% (reduced), or 0% (zero-rated)',
    'This template is free to use for your business — no sign-up required',
]
for i, note in enumerate(notes, start=13):
    ws_cover[f'B{i}'] = '•'
    ws_cover[f'B{i}'].font = Font(name='Calibri', size=10, color=CHARCOAL)
    ws_cover[f'C{i}'] = note
    ws_cover[f'C{i}'].font = Font(name='Calibri', size=10, color=CHARCOAL)

# Colour Key
ws_cover['B18'] = 'Colour Key'
ws_cover['B18'].font = Font(name='Calibri', bold=True, size=12, color=NAVY)

# Blue swatch
ws_cover['B19'].fill = INPUT_FILL
ws_cover['B19'].border = thin_border()
ws_cover['C19'] = 'You complete these — input cells'
ws_cover['C19'].font = Font(name='Calibri', size=10, color=CHARCOAL)

# Grey swatch
ws_cover['B20'].fill = READONLY_FILL
ws_cover['B20'].border = thin_border()
ws_cover['C20'] = 'Calculated automatically & locked (read-only)'
ws_cover['C20'].font = Font(name='Calibri', size=10, color=CHARCOAL)

ws_cover['B21'] = 'Table columns:'
ws_cover['B21'].font = Font(name='Calibri', bold=True, size=10, color=LABEL_GREY)
ws_cover['C21'] = 'Description / Part, Qty, Unit Price, VAT % = INPUT (blue)  |  Net, VAT, Amount = CALCULATED (grey)'
ws_cover['C21'].font = Font(name='Calibri', size=10, color=CHARCOAL)

# Promotions box
ws_cover['B23'] = 'Get more from your spreadsheets'
ws_cover['B23'].font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
ws_cover['B23'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
ws_cover.merge_cells('B23:C23')
ws_cover['B23'].alignment = ALIGN_CENTER

ws_cover['B24'] = 'OpenSheets'
ws_cover['B24'].font = Font(name='Calibri', bold=True, size=11, color=NAVY, underline='single')
ws_cover['B24'].hyperlink = 'https://opensheets.co.uk'
ws_cover['C24'] = 'Browse more free business templates'
ws_cover['C24'].font = Font(name='Calibri', size=10, color=CHARCOAL)

ws_cover['B25'] = 'Aligned.tax'
ws_cover['B25'].font = Font(name='Calibri', bold=True, size=11, color=NAVY, underline='single')
ws_cover['B25'].hyperlink = 'https://aligned.tax'
ws_cover['C25'] = 'File your MTD VAT & Income Tax returns from your spreadsheet'
ws_cover['C25'].font = Font(name='Calibri', size=10, color=CHARCOAL)

ws_cover['B26'] = 'Free to use • No sign-up required'
ws_cover['B26'].font = Font(name='Calibri', italic=True, size=9, color=MUTED)
ws_cover.merge_cells('B26:C26')

for r in range(24, 27):
    for c in ['B', 'C']:
        ws_cover[f'{c}{r}'].fill = PROMO_FILL
    ws_cover[f'B{r}'].border = Border(left=Side(style='thin', color=BORDER_GREY), bottom=Side(style='thin', color=BORDER_GREY))
    ws_cover[f'C{r}'].border = Border(right=Side(style='thin', color=BORDER_GREY), bottom=Side(style='thin', color=BORDER_GREY))

# ============================================================
# PURCHASE ORDER SHEET
# ============================================================
ws = wb.create_sheet('Purchase Order')
ws.sheet_view.showGridLines = False

# Column widths (A = margin gutter)
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 14  # party card / labels
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 12  # detail labels
ws.column_dimensions['F'].width = 12  # detail values
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 3

# --- TITLE (B2) ---
ws['B2'] = 'PURCHASE ORDER'
ws['B2'].font = TITLE_FONT
ws['B2'].alignment = ALIGN_LEFT

# --- LOGO PLACEHOLDER (G2:H4) ---
ws.merge_cells('G2:H4')
ws['G2'] = 'Add your logo here'
ws['G2'].font = Font(name='Calibri', italic=True, size=11, color=MUTED)
ws['G2'].alignment = ALIGN_CENTER
ws['G2'].border = Border(
    left=Side(style='dashed', color=BORDER_GREY),
    right=Side(style='dashed', color=BORDER_GREY),
    top=Side(style='dashed', color=BORDER_GREY),
    bottom=Side(style='dashed', color=BORDER_GREY)
)

# --- COMPANY ONE-LINER (B3) ---
ws['B3'] = '=B48&"   ·   "&B49'  # References footer name + address
ws['B3'].font = Font(name='Calibri', bold=True, size=10, color=LABEL_GREY)

# --- HEADER DIVIDER (row 4) ---
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws[f'{col}4'].border = Border(bottom=Side(style='thin', color=ROW_SEP))

# ============================================================
# TWO-COLUMN INFO SECTION (TOP-ALIGNED)
# Row 5 = labels, Row 6+ = data
# ============================================================

# --- LEFT: SUPPLIER party card ---
ws['B5'] = 'SUPPLIER'
ws['B5'].font = SECTION_LABEL_FONT

# Party card: blue fill, outer box, inner separators, italic grey prompts
party_fields = ['Supplier name', 'Address', 'Town / City', 'County', 'Postcode', 'Phone', 'Email']
party_start_row = 6
for i, field in enumerate(party_fields):
    row = party_start_row + i
    ws.merge_cells(f'B{row}:C{row}')
    cell = ws[f'B{row}']
    cell.value = field
    cell.font = PARTY_PROMPT_FONT
    cell.fill = INPUT_FILL
    cell.alignment = ALIGN_LEFT
    cell.protection = UNLOCKED
    # Outer box for first and last, separators for middle
    if i == 0:
        cell.border = Border(left=Side(style='thin', color=BORDER_GREY),
                           right=Side(style='thin', color=BORDER_GREY),
                           top=Side(style='thin', color=BORDER_GREY),
                           bottom=Side(style='thin', color=SEPARATOR_GREY))
    elif i == len(party_fields) - 1:
        cell.border = Border(left=Side(style='thin', color=BORDER_GREY),
                           right=Side(style='thin', color=BORDER_GREY),
                           top=Side(style='thin', color=SEPARATOR_GREY),
                           bottom=Side(style='thin', color=BORDER_GREY))
    else:
        cell.border = Border(left=Side(style='thin', color=BORDER_GREY),
                           right=Side(style='thin', color=BORDER_GREY),
                           top=Side(style='thin', color=SEPARATOR_GREY),
                           bottom=Side(style='thin', color=SEPARATOR_GREY))

# --- RIGHT: Detail pairs ---
ws['E5'] = 'PURCHASE ORDER DETAILS'
ws['E5'].font = SECTION_LABEL_FONT

detail_fields = [
    ('PO No:', 'PO-2026-001'),
    ('Order date:', datetime.now().strftime('%d/%m/%Y')),
    ('Required by:', ''),
    ('Buyer:', ''),
    ('Delivery terms:', ''),
]
detail_start_row = 6
for i, (label, value) in enumerate(detail_fields):
    row = detail_start_row + i
    # Label (right-aligned, merged E:D? Actually E is label, F:G merged is value)
    ws.merge_cells(f'E{row}:E{row}')
    ws[f'E{row}'] = label
    ws[f'E{row}'].font = DETAIL_LABEL_FONT
    ws[f'E{row}'].alignment = ALIGN_RIGHT
    
    # Value (merged F:G, blue input)
    ws.merge_cells(f'F{row}:G{row}')
    ws[f'F{row}'] = value
    ws[f'F{row}'].font = DETAIL_VALUE_FONT
    ws[f'F{row}'].fill = INPUT_FILL
    ws[f'F{row}'].alignment = ALIGN_LEFT
    ws[f'F{row}'].protection = UNLOCKED
    
    # Outer box for the whole detail card
    if i == 0:
        ws[f'E{row}'].border = Border(left=Side(style='thin', color=BORDER_GREY),
                                     top=Side(style='thin', color=BORDER_GREY))
        ws[f'F{row}'].border = Border(right=Side(style='thin', color=BORDER_GREY),
                                      top=Side(style='thin', color=BORDER_GREY))
    elif i == len(detail_fields) - 1:
        ws[f'E{row}'].border = Border(left=Side(style='thin', color=BORDER_GREY),
                                     bottom=Side(style='thin', color=BORDER_GREY))
        ws[f'F{row}'].border = Border(right=Side(style='thin', color=BORDER_GREY),
                                      bottom=Side(style='thin', color=BORDER_GREY))
    else:
        ws[f'E{row}'].border = Border(left=Side(style='thin', color=BORDER_GREY))
        ws[f'F{row}'].border = Border(right=Side(style='thin', color=BORDER_GREY))

# ============================================================
# COLOUR BAND (rows 14-15)
# 3 navy blocks + 1 charcoal key figure
# ============================================================
band_row = 14
band_height = 2

# Block 1: PO No
ws.merge_cells(f'B{band_row}:C{band_row+1}')
top_left = ws[f'B{band_row}']
top_left.fill = NAVY_FILL
top_left.value = 'PO NO\n=PO_NO'
top_left.font = BAND_LABEL_FONT
top_left.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Block 2: Order date
ws.merge_cells(f'D{band_row}:E{band_row+1}')
top_left = ws[f'D{band_row}']
top_left.fill = NAVY_FILL
top_left.value = 'ORDER DATE\n=ORDER_DATE'
top_left.font = BAND_LABEL_FONT
top_left.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Block 3: Required by
ws.merge_cells(f'F{band_row}:G{band_row+1}')
top_left = ws[f'F{band_row}']
top_left.fill = NAVY_FILL
top_left.value = 'REQUIRED BY\n=REQUIRED_BY'
top_left.font = BAND_LABEL_FONT
top_left.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Block 4: ORDER TOTAL (charcoal)
ws.merge_cells(f'H{band_row}:H{band_row+1}')
top_left = ws[f'H{band_row}']
top_left.fill = CHARCOAL_FILL
top_left.value = 'ORDER TOTAL\n=TOTAL'
top_left.font = BAND_LABEL_FONT
top_left.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# White vertical borders between blocks
for r in range(band_row, band_row + 2):
    ws[f'C{r}'].border = Border(right=Side(style='medium', color='FFFFFF'))
    ws[f'E{r}'].border = Border(right=Side(style='medium', color='FFFFFF'))
    ws[f'G{r}'].border = Border(right=Side(style='medium', color='FFFFFF'))

# ============================================================
# LINE-ITEMS TABLE
# Header row 17, body rows 18-23
# ============================================================
hdr_row = 17
table_cols = [
    ('B', 'Description / Part'),
    ('C', 'Qty'),
    ('D', 'Unit Price (£)'),
    ('E', 'Net (£)'),
    ('F', 'VAT %'),
    ('G', 'VAT (£)'),
    ('H', 'Amount (£)'),
]

# Header
for col, val in table_cols:
    cell = ws[f'{col}{hdr_row}']
    cell.value = val
    cell.font = TABLE_HEADER_FONT
    cell.alignment = ALIGN_CENTER if col != 'B' else ALIGN_LEFT
    cell.border = Border(bottom=Side(style='medium', color=NAVY))

# Body rows (18-23) = 6 lines
input_cols = ['B', 'C', 'D', 'F']  # Description, Qty, Unit Price, VAT %
calc_cols = ['E', 'G', 'H']  # Net, VAT, Amount

dv_vat = DataValidation(type="list", formula1='"0%,5%,20%"', allow_blank=True)

for row in range(18, 24):
    # Input columns (blue)
    for col in input_cols:
        cell = ws[f'{col}{row}']
        cell.fill = INPUT_FILL
        cell.font = TABLE_BODY_FONT
        cell.protection = UNLOCKED
        cell.border = row_sep()
        if col == 'B':
            cell.alignment = ALIGN_LEFT
        else:
            cell.alignment = ALIGN_CENTER
    
    # VAT % default
    ws[f'F{row}'] = '20%'
    dv_vat.add(f'F{row}')
    
    # Calculated columns (grey)
    # Net = IF blank, "-", Qty * Unit Price
    ws[f'E{row}'] = f'=IF(OR(C{row}="",D{row}="",C{row}=0,D{row}=0),"-",C{row}*D{row})'
    ws[f'E{row}'].fill = READONLY_FILL
    ws[f'E{row}'].font = TABLE_BODY_FONT
    ws[f'E{row}'].number_format = MONEY_FMT
    ws[f'E{row}'].alignment = ALIGN_RIGHT
    ws[f'E{row}'].border = row_sep()
    
    # VAT = IF net is "-" or VAT% blank, "-", Net * parsed_rate
    vat_formula = (
        f'=IF(OR(E{row}="-",F{row}=""),"-",'
        f'E{row}*IF(F{row}="20%",0.2,IF(F{row}="5%",0.05,0)))'
    )
    ws[f'G{row}'] = vat_formula
    ws[f'G{row}'].fill = READONLY_FILL
    ws[f'G{row}'].font = TABLE_BODY_FONT
    ws[f'G{row}'].number_format = MONEY_FMT
    ws[f'G{row}'].alignment = ALIGN_RIGHT
    ws[f'G{row}'].border = row_sep()
    
    # Amount = Net + VAT
    ws[f'H{row}'] = f'=IF(OR(E{row}="-",G{row}="-"),"-",E{row}+G{row})'
    ws[f'H{row}'].fill = READONLY_FILL
    ws[f'H{row}'].font = TABLE_BODY_FONT
    ws[f'H{row}'].number_format = MONEY_FMT
    ws[f'H{row}'].alignment = ALIGN_RIGHT
    ws[f'H{row}'].border = row_sep()

ws.add_data_validation(dv_vat)

# ============================================================
# TOTALS (row 25-27)
# ============================================================
# Subtotal
ws.merge_cells('E25:F25')
ws['E25'] = 'Subtotal'
ws['E25'].font = TOTAL_LABEL_FONT
ws['E25'].alignment = ALIGN_RIGHT
ws['E25'].border = Border(top=Side(style='thin', color=ROW_SEP))

ws['G25'] = '=SUM(E18:E23)'
ws['G25'].font = TOTAL_LABEL_FONT
ws['G25'].number_format = MONEY_FMT
ws['G25'].alignment = ALIGN_RIGHT
ws['G25'].border = Border(top=Side(style='thin', color=ROW_SEP))

ws['H25'] = '=SUM(H18:H23)'
ws['H25'].font = TOTAL_LABEL_FONT
ws['H25'].number_format = MONEY_FMT
ws['H25'].alignment = ALIGN_RIGHT
ws['H25'].border = Border(top=Side(style='thin', color=ROW_SEP))

# VAT Total
ws.merge_cells('E26:F26')
ws['E26'] = 'VAT Total'
ws['E26'].font = TOTAL_LABEL_FONT
ws['E26'].alignment = ALIGN_RIGHT

ws['G26'] = '=SUM(G18:G23)'
ws['G26'].font = TOTAL_LABEL_FONT
ws['G26'].number_format = MONEY_FMT
ws['G26'].alignment = ALIGN_RIGHT

ws['H26'] = ''

# ORDER TOTAL
ws.merge_cells('E27:F27')
ws['E27'] = 'ORDER TOTAL'
ws['E27'].font = TOTAL_AMOUNT_FONT
ws['E27'].alignment = ALIGN_RIGHT
ws['E27'].border = navy_top()

ws['G27'] = '=G25+G26'
ws['G27'].font = TOTAL_AMOUNT_FONT
ws['G27'].number_format = MONEY_FMT
ws['G27'].alignment = ALIGN_RIGHT
ws['G27'].border = navy_top()

ws['H27'] = '=H25'
ws['H27'].font = TOTAL_AMOUNT_FONT
ws['H27'].number_format = MONEY_FMT
ws['H27'].alignment = ALIGN_RIGHT
ws['H27'].border = navy_top()

# ============================================================
# NOTES + SIGNATURE (row 29-32)
# ============================================================
ws['B29'] = 'Notes / Delivery Instructions'
ws['B29'].font = SECTION_LABEL_FONT

ws.merge_cells('B30:D32')
ws['B30'] = ''
ws['B30'].fill = INPUT_FILL
ws['B30'].border = thin_border()
ws['B30'].alignment = ALIGN_TOP_LEFT
ws['B30'].protection = UNLOCKED

ws['E29'] = 'Authorised By'
ws['E29'].font = SECTION_LABEL_FONT
ws['E29'].alignment = ALIGN_RIGHT

ws.merge_cells('F30:G30')
ws['F30'] = ''
ws['F30'].border = Border(bottom=Side(style='thin', color=CHARCOAL))
ws['F30'].alignment = ALIGN_CENTER
ws['F30'].protection = UNLOCKED

ws['E31'] = 'Signature'
ws['E31'].font = NOTE_FONT
ws['E31'].alignment = ALIGN_RIGHT

# ============================================================
# FOOTER (row 34-37) — single source of truth
# ============================================================
# Top rule
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws[f'{col}34'].border = Border(top=Side(style='thin', color=ROW_SEP))

# Company name
ws['B35'] = 'YOUR COMPANY NAME LTD'
ws['B35'].font = Font(name='Calibri', bold=True, size=11, color=CHARCOAL)
ws['B35'].protection = UNLOCKED

# Address
ws['B36'] = '123 Business Street, London, EC1A 1BB'
ws['B36'].font = FOOTER_FONT
ws['B36'].protection = UNLOCKED

# Contact
ws['B37'] = 'Tel: 020 7946 0958   ·   Email: hello@yourcompany.co.uk   ·   VAT Reg: GB123456789'
ws['B37'].font = FOOTER_FONT
ws['B37'].protection = UNLOCKED

# ============================================================
# PROTECTION
# ============================================================
# Protect everything, unlock inputs
ws.protection.sheet = True
ws.protection.enable()

# ============================================================
# PRINT SETTINGS
# ============================================================
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.25  # ~18pt
ws.page_margins.right = 0.25
ws.page_margins.top = 0.33   # ~24pt
ws.page_margins.bottom = 0.33

# ============================================================
# ROW HEIGHTS (generous for print)
# ============================================================
ws.row_dimensions[2].height = 36   # Title
ws.row_dimensions[4].height = 4    # Divider
ws.row_dimensions[5].height = 18   # Section labels
for r in range(6, 13):
    ws.row_dimensions[r].height = 20  # Party/details
ws.row_dimensions[14].height = 16   # Band label
ws.row_dimensions[15].height = 28   # Band value
ws.row_dimensions[17].height = 22   # Table header
for r in range(18, 24):
    ws.row_dimensions[r].height = 22  # Table body
for r in range(25, 28):
    ws.row_dimensions[r].height = 20  # Totals
ws.row_dimensions[30].height = 24   # Notes
ws.row_dimensions[35].height = 16   # Footer

# ============================================================
# SAVE
# ============================================================
wb.save('app/public/templates/Purchase_Order_Template_v2.xlsx')
print('✅ Purchase Order Template v2 built!')
print('   • Navy house style with colour band')
print('   • Field-prompt supplier card + detail pairs')
print('   • Per-line VAT rates (0%/5%/20%)')
print('   • Blue input / grey read-only colour code')
print('   • Cover sheet with OpenSheets + Aligned.tax promos')
print('   • A4 portrait, print-ready, gridlines hidden')
