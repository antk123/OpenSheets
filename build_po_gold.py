#!/usr/bin/env python3
"""
Rebuild Purchase Order to match the gold-standard example EXACTLY.
Key fixes: ARGB color format (FF prefix), theme colors for body text,
properly escaped number formats, exact border patterns.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers, Color
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ============================================================
# PALETTE — use explicit ARGB with FF alpha prefix
# ============================================================
NAVY = 'FF1E40AF'
CHARCOAL = 'FF1F2937'
INPUT_BLUE = 'FFDBEAFE'
BORDER_GREY = 'FF9CA3AF'
SEPARATOR_GREY = 'FFCBD5E1'
ROW_SEP = 'FFE5E7EB'
MUTED = 'FF6B7280'
LABEL_GREY = 'FF374151'
WHITE = 'FFFFFFFF'

# Theme color (type='theme', value=1) — Excel default text color
THEME_TEXT = Color(theme=1, type='theme')

# ============================================================
# FONTS — Aptos Narrow, exact sizes from gold standard
# ============================================================
def make_font(name='Aptos Narrow', size=11, bold=False, italic=False, color=None):
    if color is None:
        return Font(name=name, size=size, bold=bold, italic=italic)
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

TITLE_FONT = make_font(size=28, bold=True, color=CHARCOAL)
ONE_LINER_FONT = make_font(size=11, bold=True, color=LABEL_GREY)
SECTION_LABEL_FONT = make_font(size=11, bold=True, color=MUTED)  # "SUPPLIER"
DETAIL_LABEL_FONT = make_font(size=11, bold=True, color=LABEL_GREY)  # E6 labels
DETAIL_VALUE_FONT = make_font(size=12, color=CHARCOAL)  # G6 values
PARTY_PROMPT_FONT = make_font(size=12, italic=True, color='FF9CA3AF')  # B7 placeholders
TABLE_HEADER_FONT = make_font(size=13, bold=True, color=CHARCOAL)
TABLE_BODY_FONT = make_font(size=13)  # NO color = theme default
TOTAL_LABEL_FONT = make_font(size=13, color=LABEL_GREY)  # NOT bold
TOTAL_AMOUNT_FONT = make_font(size=15, bold=True, color=CHARCOAL)
BAND_LABEL_FONT = make_font(size=9.5, bold=True, color=WHITE)
BAND_VALUE_FONT = make_font(size=16, bold=True, color=WHITE)
FOOTER_NAME_FONT = make_font(size=12, bold=True, color=CHARCOAL)
FOOTER_LINE_FONT = make_font(size=11, color=CHARCOAL)
FOOTER_CONTACT_FONT = make_font(size=11, color=MUTED)
NOTE_LABEL_FONT = make_font(size=11, bold=True, color=MUTED)
NOTE_PLACEHOLDER_FONT = make_font(size=11, italic=True, color='FF9CA3AF')
LOGO_FONT = make_font(size=10, italic=True, color='FF9CA3AF')

# ============================================================
# FILLS — explicit ARGB
# ============================================================
def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

INPUT_FILL = make_fill(INPUT_BLUE)
NAVY_FILL = make_fill(NAVY)
WHITE_FILL = make_fill(WHITE)
CHARCOAL_FILL = make_fill(CHARCOAL)

# ============================================================
# NUMBER FORMATS — match gold standard exactly with backslash escaping
# ============================================================
# The gold standard uses \£ (escaped pound sign) in the format string
# In Python we need \\£ to produce a literal backslash in the XML
MONEY_FMT = r'\£#,##0.00;\-\£#,##0.00;"–"'
QTY_FMT = r'#,##0;\-#,##0;"–"'
PCT_FMT = '0%'
DATE_FMT = 'DD/MM/YYYY'

# ============================================================
# ALIGNMENT
# ============================================================
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_TOP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ============================================================
# BORDERS — build exact patterns from gold standard
# ============================================================
def thin_side(c):
    return Side(style='thin', color=c)

def medium_side(c):
    return Side(style='medium', color=c)

# Party card: outer box + inner separators
PARTY_TOP_BORDER = Border(
    left=thin_side(BORDER_GREY), right=thin_side(BORDER_GREY),
    top=thin_side(BORDER_GREY), bottom=thin_side(SEPARATOR_GREY)
)
PARTY_MID_BORDER = Border(
    left=thin_side(BORDER_GREY), right=thin_side(BORDER_GREY),
    top=thin_side(SEPARATOR_GREY), bottom=thin_side(SEPARATOR_GREY)
)
PARTY_BOT_BORDER = Border(
    left=thin_side(BORDER_GREY), right=thin_side(BORDER_GREY),
    top=thin_side(SEPARATOR_GREY), bottom=thin_side(BORDER_GREY)
)

# Detail card: minimal borders (top/bottom only, thin)
DETAIL_TOP_BORDER = Border(
    top=thin_side(BORDER_GREY), bottom=thin_side(ROW_SEP)
)
DETAIL_MID_BORDER = Border(
    top=thin_side(ROW_SEP), bottom=thin_side(ROW_SEP)
)
DETAIL_BOT_BORDER = Border(
    top=thin_side(ROW_SEP), bottom=thin_side(BORDER_GREY)
)

# Band white vertical separators
BAND_WHITE_VERT = Border(right=medium_side(WHITE))

# Table header: thick navy bottom only
TABLE_HEADER_BORDER = Border(bottom=Side(style='thick', color=NAVY))

# Navy top border for TOTAL
NAVY_TOP_BORDER = Border(top=medium_side(NAVY))

# Notes box border
NOTES_BORDER = Border(
    left=thin_side(BORDER_GREY), right=thin_side(BORDER_GREY),
    top=thin_side(BORDER_GREY), bottom=thin_side(BORDER_GREY)
)

# Footer top separator
FOOTER_TOP_BORDER = Border(top=thin_side(ROW_SEP))

# Logo border: dashed on 3 sides, thin bottom
LOGO_BORDER = Border(
    left=Side(style='dashed', color=BORDER_GREY),
    right=Side(style='dashed', color=BORDER_GREY),
    top=Side(style='dashed', color=BORDER_GREY),
    bottom=thin_side(ROW_SEP)
)

# No border
NO_BORDER = Border()

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws.sheet_view.showGridLines = False

# ============================================================
# COLUMN WIDTHS (exact from gold standard)
# ============================================================
ws.column_dimensions['A'].width = 3.43
ws.column_dimensions['B'].width = 47.57
ws.column_dimensions['C'].width = 13.29
ws.column_dimensions['D'].width = 16.14
ws.column_dimensions['E'].width = 17.14
ws.column_dimensions['F'].width = 11.43
ws.column_dimensions['G'].width = 16.14
ws.column_dimensions['H'].width = 18.14
ws.column_dimensions['I'].width = 13.0

# ============================================================
# ROW HEIGHTS (exact from gold standard)
# ============================================================
ws.row_dimensions[2].height = 38.1
ws.row_dimensions[3].height = 18.0
for r in range(6, 14):
    ws.row_dimensions[r].height = 18.95
ws.row_dimensions[15].height = 15.95
ws.row_dimensions[16].height = 30.0
ws.row_dimensions[18].height = 24.0
for r in range(19, 31):
    ws.row_dimensions[r].height = 20.1
ws.row_dimensions[35].height = 18.0
ws.row_dimensions[36].height = 21.95
ws.row_dimensions[40].height = 18.0
ws.row_dimensions[41].height = 15.95
ws.row_dimensions[42].height = 15.95

# ============================================================
# TITLE (B2)
# ============================================================
ws['B2'] = 'Purchase Order'
ws['B2'].font = TITLE_FONT
ws['B2'].alignment = ALIGN_LEFT

# ============================================================
# LOGO PLACEHOLDER (G2:H4)
# ============================================================
ws.merge_cells('G2:H4')
ws['G2'] = 'Add your logo here'
ws['G2'].font = LOGO_FONT
ws['G2'].alignment = ALIGN_CENTER
ws['G2'].border = LOGO_BORDER

# ============================================================
# COMPANY ONE-LINER (B3)
# ============================================================
ws['B3'] = '=B40&"   ·   "&B41'
ws['B3'].font = ONE_LINER_FONT

# ============================================================
# TWO-COLUMN INFO SECTION
# Row 6 = labels, Row 7+ = data
# ============================================================

# --- LEFT: SUPPLIER party card ---
ws['B6'] = 'SUPPLIER'
ws['B6'].font = SECTION_LABEL_FONT

party_fields = ['Supplier name', 'Address', 'Town / City', 'County', 'Postcode', 'Phone', 'Email']
borders = [PARTY_TOP_BORDER, PARTY_MID_BORDER, PARTY_MID_BORDER, PARTY_MID_BORDER, 
           PARTY_MID_BORDER, PARTY_MID_BORDER, PARTY_BOT_BORDER]

for i, field in enumerate(party_fields):
    row = 7 + i
    ws.merge_cells(f'B{row}:C{row}')
    cell = ws[f'B{row}']
    cell.value = field
    cell.font = PARTY_PROMPT_FONT
    cell.fill = INPUT_FILL
    cell.alignment = ALIGN_LEFT
    cell.border = borders[i]

# --- RIGHT: Detail pairs ---
detail_fields = [
    ('PO NO.', 'PO-0001'),
    ('ORDER DATE', '=DATE(2026,6,18)'),
    ('REQUIRED BY', '=DATE(2026,7,2)'),
    ('BUYER', 'Buyer name'),
    ('DELIVERY TERMS', 'e.g. FOB / Delivered'),
]
detail_borders = [DETAIL_TOP_BORDER, DETAIL_MID_BORDER, DETAIL_MID_BORDER, 
                  DETAIL_MID_BORDER, DETAIL_BOT_BORDER]

for i, (label, value) in enumerate(detail_fields):
    row = 6 + i
    
    # Label (E, merged E:F)
    ws.merge_cells(f'E{row}:F{row}')
    ws[f'E{row}'] = label
    ws[f'E{row}'].font = DETAIL_LABEL_FONT
    ws[f'E{row}'].alignment = ALIGN_RIGHT
    ws[f'E{row}'].border = detail_borders[i]
    
    # Value (G, merged G:H)
    ws.merge_cells(f'G{row}:H{row}')
    ws[f'G{row}'] = value
    ws[f'G{row}'].font = DETAIL_VALUE_FONT
    ws[f'G{row}'].fill = INPUT_FILL
    ws[f'G{row}'].alignment = ALIGN_LEFT
    ws[f'G{row}'].border = detail_borders[i]

# ============================================================
# COLOUR BAND (rows 15-16)
# ============================================================
# Block 1: PO NO.
ws.merge_cells('B15:C15')
ws['B15'] = 'PO NO.'
ws['B15'].font = BAND_LABEL_FONT
ws['B15'].fill = NAVY_FILL
ws['B15'].alignment = Alignment(horizontal='center', vertical='bottom')

ws.merge_cells('B16:C16')
ws['B16'] = '=G6'
ws['B16'].font = BAND_VALUE_FONT
ws['B16'].fill = NAVY_FILL
ws['B16'].alignment = Alignment(horizontal='center', vertical='top')

# Block 2: ORDER DATE
ws.merge_cells('D15:E15')
ws['D15'] = 'ORDER DATE'
ws['D15'].font = BAND_LABEL_FONT
ws['D15'].fill = NAVY_FILL
ws['D15'].alignment = Alignment(horizontal='center', vertical='bottom')

ws.merge_cells('D16:E16')
ws['D16'] = '=G7'
ws['D16'].font = BAND_VALUE_FONT
ws['D16'].fill = NAVY_FILL
ws['D16'].alignment = Alignment(horizontal='center', vertical='top')

# Block 3: REQUIRED BY
ws.merge_cells('F15:G15')
ws['F15'] = 'REQUIRED BY'
ws['F15'].font = BAND_LABEL_FONT
ws['F15'].fill = NAVY_FILL
ws['F15'].alignment = Alignment(horizontal='center', vertical='bottom')

ws.merge_cells('F16:G16')
ws['F16'] = '=G8'
ws['F16'].font = BAND_VALUE_FONT
ws['F16'].fill = NAVY_FILL
ws['F16'].alignment = Alignment(horizontal='center', vertical='top')

# Block 4: ORDER TOTAL (charcoal)
ws['H15'] = 'ORDER TOTAL'
ws['H15'].font = BAND_LABEL_FONT
ws['H15'].fill = CHARCOAL_FILL
ws['H15'].alignment = Alignment(horizontal='center', vertical='bottom')

ws['H16'] = '=H33'
ws['H16'].font = BAND_VALUE_FONT
ws['H16'].fill = CHARCOAL_FILL
ws['H16'].alignment = Alignment(horizontal='center', vertical='top')
ws['H16'].number_format = r'\£#,##0;\-\£#,##0;"–"'

# White medium vertical borders between blocks
for r in range(15, 17):
    ws[f'C{r}'].border = BAND_WHITE_VERT
    ws[f'E{r}'].border = BAND_WHITE_VERT
    ws[f'G{r}'].border = BAND_WHITE_VERT

# ============================================================
# LINE-ITEMS TABLE
# Header row 18, body rows 19-30 (12 rows)
# ============================================================
headers = [
    ('B', 'Description / Part'),
    ('C', 'Qty'),
    ('D', 'Unit price'),
    ('E', 'Net'),
    ('F', 'VAT %'),
    ('G', 'VAT'),
    ('H', 'Amount'),
]

for col, val in headers:
    cell = ws[f'{col}18']
    cell.value = val
    cell.font = TABLE_HEADER_FONT
    cell.alignment = ALIGN_CENTER if col != 'B' else ALIGN_LEFT
    cell.border = TABLE_HEADER_BORDER

# Body rows (19-30) = 12 lines
# White fill, theme color text, NO borders
dv_vat = DataValidation(type="list", formula1='"0,0.05,0.2"', allow_blank=True)

for row in range(19, 31):
    # Description (B) — input, white fill, theme color
    ws[f'B{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'B{row}'].fill = WHITE_FILL
    ws[f'B{row}'].alignment = ALIGN_LEFT
    
    # Qty (C) — input, default 1, white fill, theme color
    ws[f'C{row}'] = 1
    ws[f'C{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'C{row}'].fill = WHITE_FILL
    ws[f'C{row}'].alignment = ALIGN_CENTER
    ws[f'C{row}'].number_format = QTY_FMT
    
    # Unit price (D) — input, default 1, white fill, theme color
    ws[f'D{row}'] = 1
    ws[f'D{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'D{row}'].fill = WHITE_FILL
    ws[f'D{row}'].alignment = ALIGN_CENTER
    ws[f'D{row}'].number_format = MONEY_FMT
    
    # Net (E) — calculated, white fill, theme color
    ws[f'E{row}'] = f'=IF(AND(C{row}="",D{row}=""),"",C{row}*D{row})'
    ws[f'E{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'E{row}'].fill = WHITE_FILL
    ws[f'E{row}'].alignment = ALIGN_RIGHT
    ws[f'E{row}'].number_format = MONEY_FMT
    
    # VAT % (F) — input, default: first row 0.05, rest 0.2, white fill, theme color
    ws[f'F{row}'] = 0.05 if row == 19 else 0.2
    ws[f'F{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'F{row}'].fill = WHITE_FILL
    ws[f'F{row}'].alignment = ALIGN_CENTER
    ws[f'F{row}'].number_format = PCT_FMT
    dv_vat.add(f'F{row}')
    
    # VAT (G) — calculated, white fill, theme color
    ws[f'G{row}'] = f'=IF(E{row}="","",E{row}*F{row})'
    ws[f'G{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'G{row}'].fill = WHITE_FILL
    ws[f'G{row}'].alignment = ALIGN_RIGHT
    ws[f'G{row}'].number_format = MONEY_FMT
    
    # Amount (H) — calculated, white fill, theme color
    ws[f'H{row}'] = f'=IF(E{row}="","",E{row}+G{row})'
    ws[f'H{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'H{row}'].fill = WHITE_FILL
    ws[f'H{row}'].alignment = ALIGN_RIGHT
    ws[f'H{row}'].number_format = MONEY_FMT

ws.add_data_validation(dv_vat)

# ============================================================
# TOTALS (row 31-33)
# ============================================================
# Subtotal
ws.merge_cells('F31:G31')
ws['F31'] = 'Subtotal'
ws['F31'].font = TOTAL_LABEL_FONT
ws['F31'].alignment = ALIGN_RIGHT
ws['F31'].border = Border(top=thin_side(ROW_SEP))

ws['H31'] = '=SUM(E19:E30)'
ws['H31'].font = TOTAL_LABEL_FONT
ws['H31'].fill = WHITE_FILL
ws['H31'].alignment = ALIGN_RIGHT
ws['H31'].number_format = MONEY_FMT
ws['H31'].border = Border(top=thin_side(ROW_SEP))

# VAT
ws.merge_cells('F32:G32')
ws['F32'] = 'VAT'
ws['F32'].font = TOTAL_LABEL_FONT
ws['F32'].alignment = ALIGN_RIGHT

ws['H32'] = '=SUM(G19:G30)'
ws['H32'].font = TOTAL_LABEL_FONT
ws['H32'].fill = PatternFill(start_color='FFEFF3FB', end_color='FFEFF3FB', fill_type='solid')
ws['H32'].alignment = ALIGN_RIGHT
ws['H32'].number_format = MONEY_FMT

# TOTAL
ws.merge_cells('F33:G33')
ws['F33'] = 'TOTAL'
ws['F33'].font = TOTAL_AMOUNT_FONT
ws['F33'].alignment = ALIGN_RIGHT
ws['F33'].border = NAVY_TOP_BORDER

ws['H33'] = '=H31+H32'
ws['H33'].font = TOTAL_AMOUNT_FONT
ws['H33'].alignment = ALIGN_RIGHT
ws['H33'].number_format = MONEY_FMT
ws['H33'].border = NAVY_TOP_BORDER

# ============================================================
# NOTES + SIGNATURE (row 35-36)
# ============================================================
ws['B35'] = 'NOTES'
ws['B35'].font = NOTE_LABEL_FONT

ws.merge_cells('B36:D38')
ws['B36'] = 'Delivery instructions, references, special terms…'
ws['B36'].font = NOTE_PLACEHOLDER_FONT
ws['B36'].fill = INPUT_FILL
ws['B36'].alignment = ALIGN_TOP_LEFT
ws['B36'].border = NOTES_BORDER

ws['F35'] = 'Issued by, signature:'
ws['F35'].font = Font(name='Aptos Narrow', size=11, bold=True, color=LABEL_GREY)
ws['F35'].alignment = ALIGN_RIGHT

ws['F36'].border = Border(bottom=thin_side(CHARCOAL))
ws['F36'].alignment = ALIGN_CENTER

# ============================================================
# FOOTER (row 40-42)
# ============================================================
ws['B40'] = 'Your Company Ltd'
ws['B40'].font = FOOTER_NAME_FONT
ws['B40'].border = FOOTER_TOP_BORDER

ws['B41'] = '123 Example Street, London, EC1A 1BB'
ws['B41'].font = Font(name='Aptos Narrow', size=11, color=MUTED)
ws['B41'].border = FOOTER_TOP_BORDER

ws['B42'] = 'Tel 020 0000 0000   ·   accounts@yourcompany.co.uk   ·   VAT Reg GB000000000'
ws['B42'].font = FOOTER_CONTACT_FONT
ws['B42'].border = FOOTER_TOP_BORDER

# ============================================================
# PRINT SETTINGS
# ============================================================
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25
ws.page_margins.top = 0.33
ws.page_margins.bottom = 0.33

# ============================================================
# SAVE
# ============================================================
wb.save('app/public/templates/Purchase_Order_Template_v2.xlsx')
print('✅ Purchase Order Template v2 (exact gold-standard replica) built!')
print('   • ARGB color format (FF prefix)')
print('   • Theme color for table body text')
print('   • Properly escaped number formats')
print('   • Exact border patterns from gold standard')
print('   • 28 merged cell ranges (matching gold)')
