#!/usr/bin/env python3
"""
Purchase Order — GOLD STANDARD v3
Fixes from user feedback:
- Zebra striping on calculated columns (E,G,H) — FFFFFF / FFEFF3FB alternating
- Input columns (B,C,D,F) always white FFFFFF — no zebra, "lighter"
- Band labels/values: ONLY right=medium(white) vertical separator, NO top/bottom/left borders
- Logo: complete dashed border on all 4 sides #9CA3AF
- Input cells (C,D) empty by default — no default values
- Detail value borders: L=thin,R=thin all rows; T=thin first row; B=thin last row
- Detail labels (E): NO borders at all
- VAT dropdown defaults: F19=0.05, F20-30=0.2, formatted as 0%
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# PALETTE — explicit ARGB
# ============================================================
NAVY = 'FF1E40AF'
CHARCOAL = 'FF1F2937'
INPUT_BLUE = 'FFDBEAFE'
BORDER_GREY = 'FF9CA3AF'
SEPARATOR_GREY = 'FFCBD5E1'
ROW_SEP = 'FFE5E7EB'
MUTED = 'FF6B7280'
LABEL_GREY = 'FF374151'
WHITE = 'FFFFFFFF'
ZEBRA_GREY = 'FFEFF3FB'

# ============================================================
# FONTS
# ============================================================
def f(name='Aptos Narrow', size=11, bold=False, italic=False, color=None):
    if color is None:
        return Font(name=name, size=size, bold=bold, italic=italic)
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

TITLE_FONT = f(size=28, bold=True, color=CHARCOAL)
ONE_LINER_FONT = f(size=11, bold=True, color=LABEL_GREY)
SECTION_LABEL_FONT = f(size=11, bold=True, color=MUTED)
DETAIL_LABEL_FONT = f(size=11, bold=True, color=LABEL_GREY)
DETAIL_VALUE_FONT = f(size=12, color=CHARCOAL)
PARTY_PROMPT_FONT = f(size=12, italic=True, color='FF9CA3AF')
TABLE_HEADER_FONT = f(size=13, bold=True, color=CHARCOAL)
TABLE_BODY_FONT = f(size=13)  # theme color default
TOTAL_LABEL_FONT = f(size=13, color=LABEL_GREY)
TOTAL_AMOUNT_FONT = f(size=15, bold=True, color=CHARCOAL)
BAND_LABEL_FONT = f(size=9.5, bold=True, color=WHITE)
BAND_VALUE_FONT = f(size=16, bold=True, color=WHITE)
FOOTER_NAME_FONT = f(size=12, bold=True, color=CHARCOAL)
FOOTER_CONTACT_FONT = f(size=11, color=MUTED)
NOTE_LABEL_FONT = f(size=11, bold=True, color=MUTED)
NOTE_PLACEHOLDER_FONT = f(size=11, italic=True, color='FF9CA3AF')
SIGNATURE_LABEL_FONT = f(size=11, bold=True, color=LABEL_GREY)
LOGO_FONT = f(size=10, italic=True, color='FF9CA3AF')

# ============================================================
# FILLS
# ============================================================
def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

INPUT_FILL = fill(INPUT_BLUE)
NAVY_FILL = fill(NAVY)
WHITE_FILL = fill(WHITE)
CHARCOAL_FILL = fill(CHARCOAL)
ZEBRA_FILL = fill(ZEBRA_GREY)

# ============================================================
# NUMBER FORMATS — backslash-escaped, en-dash for zero
# ============================================================
MONEY_FMT = r'\£#,##0.00;\-\£#,##0.00;"–"'
QTY_FMT = r'#,##0;\-#,##0;"–"'
PCT_FMT = '0%'
BAND_MONEY_FMT = r'\£#,##0;\-\£#,##0;"–"'

# ============================================================
# ALIGNMENT
# ============================================================
AL = Alignment(horizontal='left', vertical='center')
AR = Alignment(horizontal='right', vertical='center')
AC = Alignment(horizontal='center', vertical='center')
ATL = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ============================================================
# BORDERS — thin side helper
# ============================================================
def ts(c=BORDER_GREY):
    return Side(style='thin', color=c)

def ms(c=BORDER_GREY):
    return Side(style='medium', color=c)

def ds(c=BORDER_GREY):
    return Side(style='dashed', color=c)

# Party card borders
PARTY_TOP = Border(left=ts(), right=ts(), top=ts(), bottom=ts(SEPARATOR_GREY))
PARTY_MID = Border(left=ts(), right=ts(), top=ts(SEPARATOR_GREY), bottom=ts(SEPARATOR_GREY))
PARTY_BOT = Border(left=ts(), right=ts(), top=ts(SEPARATOR_GREY), bottom=ts())

# Detail value borders — L/R thin all rows; T thin first; B thin last
DETAIL_V_TOP = Border(left=ts(), right=ts(), top=ts(), bottom=None)
DETAIL_V_MID = Border(left=ts(), right=ts(), top=None, bottom=None)
DETAIL_V_BOT = Border(left=ts(), right=ts(), top=None, bottom=ts())

# Band — only right=medium white
BAND_SEP = Border(right=ms(WHITE))

# Logo — complete dashed box
LOGO_BORDER = Border(left=ds(), right=ds(), top=ds(), bottom=ds())

# Table header — thick navy bottom only
TABLE_HDR_BORDER = Border(bottom=Side(style='thick', color=NAVY))

# Navy top for TOTAL row
NAVY_TOP_BORDER = Border(top=Side(style='medium', color=NAVY))

# Notes box
NOTES_BORDER = Border(left=ts(), right=ts(), top=ts(), bottom=ts())

# Footer top rule
FOOTER_TOP = Border(top=ts(ROW_SEP))

# Signature line
SIGNATURE_LINE = Border(bottom=ts(CHARCOAL))

# ============================================================
# BUILD
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws.sheet_view.showGridLines = False

# Column widths
ws.column_dimensions['A'].width = 3.43
ws.column_dimensions['B'].width = 47.57
ws.column_dimensions['C'].width = 13.29
ws.column_dimensions['D'].width = 16.14
ws.column_dimensions['E'].width = 17.14
ws.column_dimensions['F'].width = 11.43
ws.column_dimensions['G'].width = 16.14
ws.column_dimensions['H'].width = 18.14
ws.column_dimensions['I'].width = 13.0

# Row heights
ws.row_dimensions[2].height = 38.1
ws.row_dimensions[3].height = 18.0
for r in range(6, 14):
    ws.row_dimensions[r].height = 18.95
ws.row_dimensions[15].height = 15.95
ws.row_dimensions[16].height = 30.0
ws.row_dimensions[18].height = 24.0
for r in range(19, 31):
    ws.row_dimensions[r].height = 20.1
ws.row_dimensions[35].height = 18.0
ws.row_dimensions[36].height = 21.95
ws.row_dimensions[40].height = 18.0
ws.row_dimensions[41].height = 15.95
ws.row_dimensions[42].height = 15.95

# --- TITLE (B2) ---
ws['B2'] = 'Purchase Order'
ws['B2'].font = TITLE_FONT
ws['B2'].alignment = AL

# --- LOGO (G2:H4) — complete dashed border ---
ws.merge_cells('G2:H4')
ws['G2'] = 'Add your logo here'
ws['G2'].font = LOGO_FONT
ws['G2'].alignment = AC
ws['G2'].border = LOGO_BORDER

# --- ONE-LINER (B3) ---
ws['B3'] = '=B40&"   ·   "&B41'
ws['B3'].font = ONE_LINER_FONT

# --- SUPPLIER label (B6) ---
ws['B6'] = 'SUPPLIER'
ws['B6'].font = SECTION_LABEL_FONT

# --- PARTY CARD (B7:C13) ---
party_fields = ['Supplier name', 'Address', 'Town / City', 'County', 'Postcode', 'Phone', 'Email']
party_borders = [PARTY_TOP, PARTY_MID, PARTY_MID, PARTY_MID, PARTY_MID, PARTY_MID, PARTY_BOT]
for i, field in enumerate(party_fields):
    row = 7 + i
    ws.merge_cells(f'B{row}:C{row}')
    c = ws[f'B{row}']
    c.value = field
    c.font = PARTY_PROMPT_FONT
    c.fill = INPUT_FILL
    c.alignment = AL
    c.border = party_borders[i]

# --- DETAIL LABELS (E6:F10) — NO borders ---
# --- DETAIL VALUES (G6:H10) — L/R thin, T first, B last ---
detail_data = [
    ('PO NO.', 'PO-0001', True, False),      # first row
    ('ORDER DATE', '=DATE(2026,6,18)', False, False),
    ('REQUIRED BY', '=DATE(2026,7,2)', False, False),
    ('BUYER', 'Buyer name', False, False),
    ('DELIVERY TERMS', 'e.g. FOB / Delivered', False, True),  # last row
]
for i, (label, value, is_first, is_last) in enumerate(detail_data):
    row = 6 + i
    # Label — merged E:F, no borders
    ws.merge_cells(f'E{row}:F{row}')
    ws[f'E{row}'] = label
    ws[f'E{row}'].font = DETAIL_LABEL_FONT
    ws[f'E{row}'].alignment = AR
    ws[f'E{row}'].border = Border()  # NO borders

    # Value — merged G:H, blue fill, L/R thin
    ws.merge_cells(f'G{row}:H{row}')
    ws[f'G{row}'] = value
    ws[f'G{row}'].font = DETAIL_VALUE_FONT
    ws[f'G{row}'].fill = INPUT_FILL
    ws[f'G{row}'].alignment = AL
    if is_first:
        ws[f'G{row}'].border = DETAIL_V_TOP
    elif is_last:
        ws[f'G{row}'].border = DETAIL_V_BOT
    else:
        ws[f'G{row}'].border = DETAIL_V_MID

# --- COLOUR BAND (rows 15-16) ---
# Block 1: PO NO.
ws.merge_cells('B15:C15')
ws['B15'] = 'PO NO.'
ws['B15'].font = BAND_LABEL_FONT
ws['B15'].fill = NAVY_FILL
ws['B15'].alignment = AC
ws['B15'].border = BAND_SEP

ws.merge_cells('B16:C16')
ws['B16'] = '=G6'
ws['B16'].font = BAND_VALUE_FONT
ws['B16'].fill = NAVY_FILL
ws['B16'].alignment = AC
ws['B16'].border = BAND_SEP

# Block 2: ORDER DATE
ws.merge_cells('D15:E15')
ws['D15'] = 'ORDER DATE'
ws['D15'].font = BAND_LABEL_FONT
ws['D15'].fill = NAVY_FILL
ws['D15'].alignment = AC
ws['D15'].border = BAND_SEP

ws.merge_cells('D16:E16')
ws['D16'] = '=G7'
ws['D16'].font = BAND_VALUE_FONT
ws['D16'].fill = NAVY_FILL
ws['D16'].alignment = AC
ws['D16'].border = BAND_SEP

# Block 3: REQUIRED BY
ws.merge_cells('F15:G15')
ws['F15'] = 'REQUIRED BY'
ws['F15'].font = BAND_LABEL_FONT
ws['F15'].fill = NAVY_FILL
ws['F15'].alignment = AC
ws['F15'].border = BAND_SEP

ws.merge_cells('F16:G16')
ws['F16'] = '=G8'
ws['F16'].font = BAND_VALUE_FONT
ws['F16'].fill = NAVY_FILL
ws['F16'].alignment = AC
ws['F16'].border = BAND_SEP

# Block 4: ORDER TOTAL (charcoal)
ws['H15'] = 'ORDER TOTAL'
ws['H15'].font = BAND_LABEL_FONT
ws['H15'].fill = CHARCOAL_FILL
ws['H15'].alignment = AC
ws['H15'].border = Border()  # no borders on last block

ws['H16'] = '=H33'
ws['H16'].font = BAND_VALUE_FONT
ws['H16'].fill = CHARCOAL_FILL
ws['H16'].alignment = AC
ws['H16'].number_format = BAND_MONEY_FMT
ws['H16'].border = Border()  # no borders on last block

# --- TABLE HEADER (row 18) ---
headers = [
    ('B', 'Description / Part', AL),
    ('C', 'Qty', AC),
    ('D', 'Unit price', AC),
    ('E', 'Net', AR),
    ('F', 'VAT %', AC),
    ('G', 'VAT', AR),
    ('H', 'Amount', AR),
]
for col, val, align in headers:
    c = ws[f'{col}18']
    c.value = val
    c.font = TABLE_HEADER_FONT
    c.alignment = align
    c.border = TABLE_HDR_BORDER

# --- TABLE BODY (rows 19-30) ---
# Zebra: calculated columns (E,G,H) alternate FFFFFF / FFEFF3FB
# Input columns (B,C,D,F) always FFFFFF (white, no zebra)
dv_vat = DataValidation(type="list", formula1='"0,0.05,0.2"', allow_blank=True)

for row in range(19, 31):
    # Determine zebra fill for this row
    is_even = (row % 2 == 0)
    calc_fill = ZEBRA_FILL if is_even else WHITE_FILL

    # B — Description (input, always white, no zebra)
    ws[f'B{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'B{row}'].fill = WHITE_FILL
    ws[f'B{row}'].alignment = AL

    # C — Qty (input, always white, EMPTY default)
    ws[f'C{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'C{row}'].fill = WHITE_FILL
    ws[f'C{row}'].alignment = AC
    ws[f'C{row}'].number_format = QTY_FMT
    # EMPTY — no default value

    # D — Unit price (input, always white, EMPTY default)
    ws[f'D{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'D{row}'].fill = WHITE_FILL
    ws[f'D{row}'].alignment = AC
    ws[f'D{row}'].number_format = MONEY_FMT
    # EMPTY — no default value

    # E — Net (calculated, zebra fill)
    ws[f'E{row}'] = f'=IF(AND(C{row}="",D{row}=""),"",C{row}*D{row})'
    ws[f'E{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'E{row}'].fill = calc_fill
    ws[f'E{row}'].alignment = AR
    ws[f'E{row}'].number_format = MONEY_FMT

    # F — VAT % (input, always white, dropdown)
    ws[f'F{row}'] = 0.05 if row == 19 else 0.2
    ws[f'F{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'F{row}'].fill = WHITE_FILL
    ws[f'F{row}'].alignment = AC
    ws[f'F{row}'].number_format = PCT_FMT
    dv_vat.add(f'F{row}')

    # G — VAT (calculated, zebra fill)
    ws[f'G{row}'] = f'=IF(E{row}="","",E{row}*F{row})'
    ws[f'G{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'G{row}'].fill = calc_fill
    ws[f'G{row}'].alignment = AR
    ws[f'G{row}'].number_format = MONEY_FMT

    # H — Amount (calculated, zebra fill)
    ws[f'H{row}'] = f'=IF(E{row}="","",E{row}+G{row})'
    ws[f'H{row}'].font = Font(name='Aptos Narrow', size=13, color=Color(theme=1, type='theme'))
    ws[f'H{row}'].fill = calc_fill
    ws[f'H{row}'].alignment = AR
    ws[f'H{row}'].number_format = MONEY_FMT

ws.add_data_validation(dv_vat)

# --- TOTALS (row 31-33) ---
# Subtotal
ws.merge_cells('F31:G31')
ws['F31'] = 'Subtotal'
ws['F31'].font = TOTAL_LABEL_FONT
ws['F31'].alignment = AR
ws['F31'].border = Border(top=ts(ROW_SEP))

ws['H31'] = '=SUM(E19:E30)'
ws['H31'].font = TOTAL_LABEL_FONT
ws['H31'].alignment = AR
ws['H31'].number_format = MONEY_FMT
ws['H31'].border = Border(top=ts(ROW_SEP))

# VAT
ws.merge_cells('F32:G32')
ws['F32'] = 'VAT'
ws['F32'].font = TOTAL_LABEL_FONT
ws['F32'].alignment = AR

ws['H32'] = '=SUM(G19:G30)'
ws['H32'].font = TOTAL_LABEL_FONT
ws['H32'].fill = ZEBRA_FILL  # FFEFF3FB like gold standard
ws['H32'].alignment = AR
ws['H32'].number_format = MONEY_FMT

# TOTAL
ws.merge_cells('F33:G33')
ws['F33'] = 'TOTAL'
ws['F33'].font = TOTAL_AMOUNT_FONT
ws['F33'].alignment = AR
ws['F33'].border = NAVY_TOP_BORDER

ws['H33'] = '=H31+H32'
ws['H33'].font = TOTAL_AMOUNT_FONT
ws['H33'].alignment = AR
ws['H33'].number_format = MONEY_FMT
ws['H33'].border = NAVY_TOP_BORDER

# --- NOTES + SIGNATURE (row 35-36) ---
ws['B35'] = 'NOTES'
ws['B35'].font = NOTE_LABEL_FONT

ws.merge_cells('B36:D38')
ws['B36'] = 'Delivery instructions, references, special terms…'
ws['B36'].font = NOTE_PLACEHOLDER_FONT
ws['B36'].fill = INPUT_FILL
ws['B36'].alignment = ATL
ws['B36'].border = NOTES_BORDER

ws['F35'] = 'Issued by, signature:'
ws['F35'].font = SIGNATURE_LABEL_FONT
ws['F35'].alignment = AR

# Signature line — single cell F36, NO merge
ws['F36'].border = SIGNATURE_LINE
ws['F36'].alignment = AC

# --- FOOTER (row 40-42) ---
ws['B40'] = 'Your Company Ltd'
ws['B40'].font = FOOTER_NAME_FONT
ws['B40'].border = FOOTER_TOP

ws['B41'] = '123 Example Street, London, EC1A 1BB'
ws['B41'].font = f(size=11, color=MUTED)
ws['B41'].border = FOOTER_TOP

ws['B42'] = 'Tel 020 0000 0000   ·   accounts@yourcompany.co.uk   ·   VAT Reg GB000000000'
ws['B42'].font = FOOTER_CONTACT_FONT
ws['B42'].border = FOOTER_TOP

# --- PRINT ---
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25
ws.page_margins.top = 0.33
ws.page_margins.bottom = 0.33

# --- SAVE ---
wb.save('app/public/templates/Purchase_Order_Template_v2.xlsx')
print('✅ PO v3 built!')
print('   • Zebra on calculated cols (E,G,H) — FFFFFF / FFEFF3FB')
print('   • Input cols (B,C,D,F) always white — no zebra, lighter')
print('   • Band: only right=medium(white) separator, no top/bottom/left')
print('   • Logo: complete dashed border on all 4 sides')
print('   • Input cells (C,D) empty by default')
print('   • Detail values: L/R thin; T first row; B last row')
print('   • Detail labels: NO borders')
print('   • VAT dropdown: F19=5%, F20-30=20%')
