#!/usr/bin/env python3
"""
Purchase Order — v4
User corrections:
- Delivery Terms dropdown
- Party card: complete border ALL 4 sides on every row (thin #9CA3AF)
- Logo: dashed L/R/T + thin B (matching gold standard)
- Detail values: complete outer box (L/R all rows, T first, B last)
- Table body: BLUE fill (#DBEAFE) for input cols (B,C,D,F), WHITE for calculated (E,G,H) — NO zebra
- Input cells EMPTY by default (no demo values)
- Sheet protection at end
- Cover sheet (ALWAYS included per skill)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ============================================================
# PALETTE
# ============================================================
NAVY = 'FF1E40AF'
CHARCOAL = 'FF1F2937'
INPUT_BLUE = 'FFDBEAFE'
BORDER_GREY = 'FF9CA3AF'
SEPARATOR_GREY = 'FFCBD5E1'
ROW_SEP = 'FFE5E7EB'
MUTED = 'FF6B7280'
LABEL_GREY = 'FF374151'
WHITE = 'FFFFFFFF'
PALE_BLUE = 'FFEFF6FF'

# ============================================================
# FONTS
# ============================================================
def make_font(name='Aptos Narrow', size=11, bold=False, italic=False, color=None, underline=None):
    kwargs = dict(name=name, size=size, bold=bold, italic=italic)
    if color is not None:
        kwargs['color'] = color
    if underline is not None:
        kwargs['underline'] = underline
    return Font(**kwargs)

TITLE_FONT = make_font(size=28, bold=True, color=CHARCOAL)
ONE_LINER_FONT = make_font(size=11, bold=True, color=LABEL_GREY)
SECTION_LABEL_FONT = make_font(size=11, bold=True, color=MUTED)
DETAIL_LABEL_FONT = make_font(size=11, bold=True, color=LABEL_GREY)
DETAIL_VALUE_FONT = make_font(size=12, color=CHARCOAL)
PARTY_PROMPT_FONT = make_font(size=12, italic=True, color='FF9CA3AF')
TABLE_HEADER_FONT = make_font(size=13, bold=True, color=CHARCOAL)
TABLE_BODY_FONT = make_font(size=13)  # theme color default
TOTAL_LABEL_FONT = make_font(size=13, color=LABEL_GREY)
TOTAL_AMOUNT_FONT = make_font(size=15, bold=True, color=CHARCOAL)
BAND_LABEL_FONT = make_font(size=9.5, bold=True, color=WHITE)
BAND_VALUE_FONT = make_font(size=16, bold=True, color=WHITE)
FOOTER_NAME_FONT = make_font(size=12, bold=True, color=CHARCOAL)
FOOTER_CONTACT_FONT = make_font(size=11, color=MUTED)
NOTE_LABEL_FONT = make_font(size=11, bold=True, color=MUTED)
NOTE_PLACEHOLDER_FONT = make_font(size=11, italic=True, color='FF9CA3AF')
SIGNATURE_LABEL_FONT = make_font(size=11, bold=True, color=LABEL_GREY)
LOGO_FONT = make_font(size=10, italic=True, color='FF9CA3AF')
COVER_TITLE_FONT = make_font(size=24, bold=True, color=CHARCOAL)
COVER_SECTION_FONT = make_font(size=13, bold=True, color=CHARCOAL)
COVER_BODY_FONT = make_font(size=11, color=CHARCOAL)
COVER_LINK_FONT = make_font(size=11, bold=True, color=NAVY, underline='single')

# ============================================================
# FILLS
# ============================================================
def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

INPUT_FILL = make_fill(INPUT_BLUE)
NAVY_FILL = make_fill(NAVY)
WHITE_FILL = make_fill(WHITE)
CHARCOAL_FILL = make_fill(CHARCOAL)
PALE_BLUE_FILL = make_fill(PALE_BLUE)

# ============================================================
# NUMBER FORMATS
# ============================================================
MONEY_FMT = r'\£#,##0.00;\-\£#,##0.00;"–"'
QTY_FMT = r'#,##0;\-#,##0;"–"'
PCT_FMT = '0%'
BAND_MONEY_FMT = r'\£#,##0;\-\£#,##0;"–"'
DATE_FMT = 'DD/MM/YYYY'

# ============================================================
# ALIGNMENT
# ============================================================
AL = Alignment(horizontal='left', vertical='center')
AR = Alignment(horizontal='right', vertical='center')
AC = Alignment(horizontal='center', vertical='center')
ATL = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ============================================================
# BORDERS
# ============================================================
def thin_side(c=BORDER_GREY):
    return Side(style='thin', color=c)

def dashed_side(c=BORDER_GREY):
    return Side(style='dashed', color=c)

def medium_side(c=BORDER_GREY):
    return Side(style='medium', color=c)

# Party card — ALL rows have thin border on ALL 4 sides (same color)
PARTY_BORDER = Border(left=thin_side(), right=thin_side(), top=thin_side(), bottom=thin_side())

# Detail values — L/R thin all rows; T thin first; B thin last
DETAIL_V_TOP = Border(left=thin_side(), right=thin_side(), top=thin_side(), bottom=None)
DETAIL_V_MID = Border(left=thin_side(), right=thin_side(), top=None, bottom=None)
DETAIL_V_BOT = Border(left=thin_side(), right=thin_side(), top=None, bottom=thin_side())

# Logo — dashed L/R/T + thin B (matching gold standard)
LOGO_BORDER = Border(left=dashed_side(), right=dashed_side(), top=dashed_side(), bottom=thin_side())

# Table header — thick navy bottom only
TABLE_HDR_BORDER = Border(bottom=Side(style='thick', color=NAVY))

# Band separator — medium white vertical
BAND_SEP = Border(right=medium_side(WHITE))

# Notes box
NOTES_BORDER = Border(left=thin_side(), right=thin_side(), top=thin_side(), bottom=thin_side())

# Signature line
SIGNATURE_LINE = Border(bottom=thin_side(CHARCOAL))

# Footer top rule
FOOTER_TOP = Border(top=thin_side(ROW_SEP))

# Navy top for TOTAL row
NAVY_TOP_BORDER = Border(top=Side(style='medium', color=NAVY))

# Row separator for totals
TOTAL_SEP = Border(top=thin_side(ROW_SEP))

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# ============================================================
# PO SHEET (Sheet1)
# ============================================================
ws = wb.active
ws.title = 'Sheet1'
ws.sheet_view.showGridLines = False

# Column widths
ws.column_dimensions['A'].width = 3.43
ws.column_dimensions['B'].width = 47.57
ws.column_dimensions['C'].width = 13.29
ws.column_dimensions['D'].width = 16.14
ws.column_dimensions['E'].width = 17.14
ws.column_dimensions['F'].width = 11.43
ws.column_dimensions['G'].width = 16.14
ws.column_dimensions['H'].width = 18.14
ws.column_dimensions['I'].width = 13.0

# Row heights
ws.row_dimensions[2].height = 38.1
ws.row_dimensions[3].height = 18.0
for r in range(6, 14):
    ws.row_dimensions[r].height = 18.95
ws.row_dimensions[15].height = 15.95
ws.row_dimensions[16].height = 30.0
ws.row_dimensions[18].height = 24.0
for r in range(19, 31):
    ws.row_dimensions[r].height = 20.1
ws.row_dimensions[35].height = 18.0
ws.row_dimensions[36].height = 21.95
ws.row_dimensions[40].height = 18.0
ws.row_dimensions[41].height = 15.95
ws.row_dimensions[42].height = 15.95

# --- TITLE (B2) ---
ws['B2'] = 'Purchase Order'
ws['B2'].font = TITLE_FONT
ws['B2'].alignment = AL

# --- LOGO (G2:H4) ---
ws.merge_cells('G2:H4')
ws['G2'] = 'Add your logo here'
ws['G2'].font = LOGO_FONT
ws['G2'].alignment = AC
ws['G2'].border = LOGO_BORDER

# --- ONE-LINER (B3) ---
ws['B3'] = '=B40&"   ·   "&B41'
ws['B3'].font = ONE_LINER_FONT

# --- SUPPLIER label (B6) ---
ws['B6'] = 'SUPPLIER'
ws['B6'].font = SECTION_LABEL_FONT

# --- PARTY CARD (B7:C13) — complete border ALL 4 sides on every row ---
party_fields = ['Supplier name', 'Address', 'Town / City', 'County', 'Postcode', 'Phone', 'Email']
for i, field in enumerate(party_fields):
    row = 7 + i
    ws.merge_cells(f'B{row}:C{row}')
    c = ws[f'B{row}']
    c.value = field
    c.font = PARTY_PROMPT_FONT
    c.fill = INPUT_FILL
    c.alignment = AL
    c.border = PARTY_BORDER

# --- DETAIL LABELS (E6:F10) — no borders, right-aligned ---
detail_labels = ['PO NO.', 'ORDER DATE', 'REQUIRED BY', 'BUYER', 'DELIVERY TERMS']
for i, label in enumerate(detail_labels):
    row = 6 + i
    ws.merge_cells(f'E{row}:F{row}')
    ws[f'E{row}'] = label
    ws[f'E{row}'].font = DETAIL_LABEL_FONT
    ws[f'E{row}'].alignment = AR

# --- DETAIL VALUES (G6:H10) — blue fill, complete outer box ---
detail_values = ['PO-0001', '=DATE(2026,6,18)', '=DATE(2026,7,2)', 'Buyer name', 'Delivered']
detail_borders = [DETAIL_V_TOP, DETAIL_V_MID, DETAIL_V_MID, DETAIL_V_MID, DETAIL_V_BOT]
for i, value in enumerate(detail_values):
    row = 6 + i
    ws.merge_cells(f'G{row}:H{row}')
    ws[f'G{row}'] = value
    ws[f'G{row}'].font = DETAIL_VALUE_FONT
    ws[f'G{row}'].fill = INPUT_FILL
    ws[f'G{row}'].alignment = AL
    ws[f'G{row}'].border = detail_borders[i]

# Delivery Terms dropdown
dv_terms = DataValidation(type="list", formula1='"FOB,Delivered,Ex Works,CIF,DDP,DAP,CFR"', allow_blank=True)
ws.add_data_validation(dv_terms)
dv_terms.add('G10')

# --- COLOUR BAND (rows 15-16) ---
# Block 1: PO NO.
ws.merge_cells('B15:C15')
ws['B15'] = 'PO NO.'
ws['B15'].font = BAND_LABEL_FONT
ws['B15'].fill = NAVY_FILL
ws['B15'].alignment = AC
ws['B15'].border = BAND_SEP

ws.merge_cells('B16:C16')
ws['B16'] = '=G6'
ws['B16'].font = BAND_VALUE_FONT
ws['B16'].fill = NAVY_FILL
ws['B16'].alignment = AC
ws['B16'].border = BAND_SEP

# Block 2: ORDER DATE
ws.merge_cells('D15:E15')
ws['D15'] = 'ORDER DATE'
ws['D15'].font = BAND_LABEL_FONT
ws['D15'].fill = NAVY_FILL
ws['D15'].alignment = AC
ws['D15'].border = BAND_SEP

ws.merge_cells('D16:E16')
ws['D16'] = '=G7'
ws['D16'].font = BAND_VALUE_FONT
ws['D16'].fill = NAVY_FILL
ws['D16'].alignment = AC
ws['D16'].border = BAND_SEP

# Block 3: REQUIRED BY
ws.merge_cells('F15:G15')
ws['F15'] = 'REQUIRED BY'
ws['F15'].font = BAND_LABEL_FONT
ws['F15'].fill = NAVY_FILL
ws['F15'].alignment = AC
ws['F15'].border = BAND_SEP

ws.merge_cells('F16:G16')
ws['F16'] = '=G8'
ws['F16'].font = BAND_VALUE_FONT
ws['F16'].fill = NAVY_FILL
ws['F16'].alignment = AC
ws['F16'].border = BAND_SEP

# Block 4: ORDER TOTAL (charcoal)
ws['H15'] = 'ORDER TOTAL'
ws['H15'].font = BAND_LABEL_FONT
ws['H15'].fill = CHARCOAL_FILL
ws['H15'].alignment = AC

ws['H16'] = '=H33'
ws['H16'].font = BAND_VALUE_FONT
ws['H16'].fill = CHARCOAL_FILL
ws['H16'].alignment = AC
ws['H16'].number_format = BAND_MONEY_FMT

# --- TABLE HEADER (row 18) ---
headers = [
    ('B', 'Description / Part', AL),
    ('C', 'Qty', AC),
    ('D', 'Unit price', AC),
    ('E', 'Net', AR),
    ('F', 'VAT %', AC),
    ('G', 'VAT', AR),
    ('H', 'Amount', AR),
]
for col, val, align in headers:
    c = ws[f'{col}18']
    c.value = val
    c.font = TABLE_HEADER_FONT
    c.alignment = align
    c.border = TABLE_HDR_BORDER

# --- TABLE BODY (rows 19-30) ---
# BLUE fill for input cols (B, C, D, F)
# WHITE/default for calculated cols (E, G, H)
# EMPTY defaults for input cells (B, C, D)
# VAT dropdown with default values
dv_vat = DataValidation(type="list", formula1='"0,0.05,0.2"', allow_blank=True)

for row in range(19, 31):
    # B — Description (input, BLUE)
    ws[f'B{row}'].font = TABLE_BODY_FONT
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].alignment = AL
    # EMPTY — no default value
    
    # C — Qty (input, BLUE, EMPTY)
    ws[f'C{row}'].font = TABLE_BODY_FONT
    ws[f'C{row}'].fill = INPUT_FILL
    ws[f'C{row}'].alignment = AC
    ws[f'C{row}'].number_format = QTY_FMT
    # EMPTY — no default value
    
    # D — Unit price (input, BLUE, EMPTY)
    ws[f'D{row}'].font = TABLE_BODY_FONT
    ws[f'D{row}'].fill = INPUT_FILL
    ws[f'D{row}'].alignment = AC
    ws[f'D{row}'].number_format = MONEY_FMT
    # EMPTY — no default value
    
    # E — Net (calculated, WHITE/default)
    ws[f'E{row}'] = f'=IF(AND(C{row}="",D{row}=""),"",C{row}*D{row})'
    ws[f'E{row}'].font = TABLE_BODY_FONT
    ws[f'E{row}'].alignment = AR
    ws[f'E{row}'].number_format = MONEY_FMT
    
    # F — VAT % (input, BLUE, dropdown)
    ws[f'F{row}'] = 0.05 if row == 19 else 0.2
    ws[f'F{row}'].font = TABLE_BODY_FONT
    ws[f'F{row}'].fill = INPUT_FILL
    ws[f'F{row}'].alignment = AC
    ws[f'F{row}'].number_format = PCT_FMT
    dv_vat.add(f'F{row}')
    
    # G — VAT (calculated, WHITE/default)
    ws[f'G{row}'] = f'=IF(E{row}="","",E{row}*F{row})'
    ws[f'G{row}'].font = TABLE_BODY_FONT
    ws[f'G{row}'].alignment = AR
    ws[f'G{row}'].number_format = MONEY_FMT
    
    # H — Amount (calculated, WHITE/default)
    ws[f'H{row}'] = f'=IF(E{row}="","",E{row}+G{row})'
    ws[f'H{row}'].font = TABLE_BODY_FONT
    ws[f'H{row}'].alignment = AR
    ws[f'H{row}'].number_format = MONEY_FMT

ws.add_data_validation(dv_vat)

# --- TOTALS (row 31-33) ---
# Subtotal
ws.merge_cells('F31:G31')
ws['F31'] = 'Subtotal'
ws['F31'].font = TOTAL_LABEL_FONT
ws['F31'].alignment = AR
ws['F31'].border = TOTAL_SEP

ws['H31'] = '=SUM(E19:E30)'
ws['H31'].font = TOTAL_LABEL_FONT
ws['H31'].alignment = AR
ws['H31'].number_format = MONEY_FMT
ws['H31'].border = TOTAL_SEP

# VAT
ws.merge_cells('F32:G32')
ws['F32'] = 'VAT'
ws['F32'].font = TOTAL_LABEL_FONT
ws['F32'].alignment = AR

ws['H32'] = '=SUM(G19:G30)'
ws['H32'].font = TOTAL_LABEL_FONT
ws['H32'].alignment = AR
ws['H32'].number_format = MONEY_FMT

# TOTAL
ws.merge_cells('F33:G33')
ws['F33'] = 'TOTAL'
ws['F33'].font = TOTAL_AMOUNT_FONT
ws['F33'].alignment = AR
ws['F33'].border = NAVY_TOP_BORDER

ws['H33'] = '=H31+H32'
ws['H33'].font = TOTAL_AMOUNT_FONT
ws['H33'].alignment = AR
ws['H33'].number_format = MONEY_FMT
ws['H33'].border = NAVY_TOP_BORDER

# --- NOTES + SIGNATURE (row 35-36) ---
ws['B35'] = 'NOTES'
ws['B35'].font = NOTE_LABEL_FONT

ws.merge_cells('B36:D38')
ws['B36'] = 'Delivery instructions, references, special terms…'
ws['B36'].font = NOTE_PLACEHOLDER_FONT
ws['B36'].fill = INPUT_FILL
ws['B36'].alignment = ATL
ws['B36'].border = NOTES_BORDER

ws['F35'] = 'Issued by, signature:'
ws['F35'].font = SIGNATURE_LABEL_FONT
ws['F35'].alignment = AR

ws['F36'].border = SIGNATURE_LINE
ws['F36'].alignment = AC

# --- FOOTER (row 40-42) ---
ws['B40'] = 'Your Company Ltd'
ws['B40'].font = FOOTER_NAME_FONT
ws['B40'].border = FOOTER_TOP

ws['B41'] = '123 Example Street, London, EC1A 1BB'
ws['B41'].font = make_font(size=11, color=MUTED)
ws['B41'].border = FOOTER_TOP

ws['B42'] = 'Tel 020 0000 0000   ·   accounts@yourcompany.co.uk   ·   VAT Reg GB000000000'
ws['B42'].font = FOOTER_CONTACT_FONT
ws['B42'].border = FOOTER_TOP

# --- PRINT SETUP ---
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25
ws.page_margins.top = 0.33
ws.page_margins.bottom = 0.33

# ============================================================
# SHEET PROTECTION
# ============================================================
# Unlock input cells
# Logo
ws['G2'].protection = Protection(locked=False)

# Party card
for row in range(7, 14):
    ws[f'B{row}'].protection = Protection(locked=False)
    ws[f'C{row}'].protection = Protection(locked=False)

# Detail values
for row in range(6, 11):
    ws[f'G{row}'].protection = Protection(locked=False)
    ws[f'H{row}'].protection = Protection(locked=False)

# Table input cells (B, C, D, F)
for row in range(19, 31):
    for col in ['B', 'C', 'D', 'F']:
        ws[f'{col}{row}'].protection = Protection(locked=False)

# Notes
for row in range(36, 39):
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].protection = Protection(locked=False)

# Footer
for row in range(40, 43):
    ws[f'B{row}'].protection = Protection(locked=False)

# Protect sheet
ws.protection.sheet = True
ws.protection.objects = True
ws.protection.insertHyperlinks = True
ws.protection.selectLockedCells = True
ws.protection.selectUnlockedCells = True
ws.protection.formatCells = False
ws.protection.formatColumns = False
ws.protection.formatRows = False
ws.protection.insertColumns = False
ws.protection.insertRows = False
ws.protection.deleteColumns = False
ws.protection.deleteRows = False
ws.protection.sort = False
ws.protection.autoFilter = False
ws.protection.pivotTables = False
ws.protection.scenarios = False

# ============================================================
# COVER SHEET
# ============================================================
ws_cover = wb.create_sheet('Cover', 0)  # Insert as first sheet
ws_cover.sheet_view.showGridLines = False

# Column widths for cover
ws_cover.column_dimensions['A'].width = 3.43
ws_cover.column_dimensions['B'].width = 60
ws_cover.column_dimensions['C'].width = 20
ws_cover.column_dimensions['D'].width = 20
ws_cover.column_dimensions['E'].width = 20

# Title
ws_cover['B2'] = 'Purchase Order TEMPLATE'
ws_cover['B2'].font = COVER_TITLE_FONT
ws_cover['B2'].alignment = AL

ws_cover['B3'] = 'Professional purchase order for supplier orders'
ws_cover['B3'].font = COVER_BODY_FONT
ws_cover['B3'].alignment = AL

# How to Use
ws_cover['B5'] = 'How to Use'
ws_cover['B5'].font = COVER_SECTION_FONT
ws_cover['B5'].border = Border(bottom=Side(style='thick', color=NAVY))

how_to_use = [
    '1. Enter your company details in the footer (company name, address, contact, VAT number)',
    '2. Fill in the supplier information in the left-hand card (name, address, contact details)',
    '3. Enter the PO details in the right-hand card (PO number, order date, required-by date, buyer, delivery terms)',
    '4. Add line items in the table — enter description, quantity, unit price, and select the VAT rate',
    '5. The net, VAT, and amount columns calculate automatically',
    '6. Add any notes or delivery instructions in the notes box',
    '7. Print or save as PDF when ready',
]
for i, text in enumerate(how_to_use):
    row = 6 + i
    ws_cover[f'B{row}'] = text
    ws_cover[f'B{row}'].font = COVER_BODY_FONT
    ws_cover[f'B{row}'].alignment = ATL

# Notes
ws_cover['B14'] = 'Notes'
ws_cover['B14'].font = COVER_SECTION_FONT
ws_cover['B14'].border = Border(bottom=Side(style='thick', color=NAVY))

notes = [
    '• Blue cells are for your input — totals calculate automatically',
    '• VAT rates are per line — choose 0%, 5%, or 20% from the dropdown',
    '• Order total updates automatically from the line items',
    '• Only blue cells can be edited — calculated cells are locked',
]
for i, text in enumerate(notes):
    row = 15 + i
    ws_cover[f'B{row}'] = text
    ws_cover[f'B{row}'].font = COVER_BODY_FONT
    ws_cover[f'B{row}'].alignment = ATL

# Key / Colour Legend
ws_cover['B20'] = 'Key'
ws_cover['B20'].font = COVER_SECTION_FONT
ws_cover['B20'].border = Border(bottom=Side(style='thick', color=NAVY))

# Blue swatch
ws_cover.merge_cells('B21:C21')
ws_cover['B21'].fill = INPUT_FILL
ws_cover['B21'].border = Border(left=thin_side(), right=thin_side(), top=thin_side(), bottom=thin_side())
ws_cover['D21'] = 'You complete these'
ws_cover['D21'].font = COVER_BODY_FONT
ws_cover['D21'].alignment = AL

# White swatch
ws_cover.merge_cells('B22:C22')
ws_cover['B22'].fill = WHITE_FILL
ws_cover['B22'].border = Border(left=thin_side(), right=thin_side(), top=thin_side(), bottom=thin_side())
ws_cover['D22'] = 'Calculated automatically & locked (read-only)'
ws_cover['D22'].font = COVER_BODY_FONT
ws_cover['D22'].alignment = AL

# Column legend
ws_cover['B24'] = 'Table columns: B-D (Description, Qty, Unit price) = input | E (Net) = calculated | F (VAT %) = input | G-H (VAT, Amount) = calculated'
ws_cover['B24'].font = make_font(size=10, italic=True, color=MUTED)
ws_cover['B24'].alignment = ATL

# Promotions box
ws_cover['B26'] = 'Get more from your spreadsheets'
ws_cover['B26'].font = make_font(size=12, bold=True, color=WHITE)
ws_cover['B26'].fill = NAVY_FILL
ws_cover['B26'].alignment = AL
ws_cover.merge_cells('B26:E26')

ws_cover['B27'] = 'OpenSheets — browse more free business templates'
ws_cover['B27'].font = COVER_LINK_FONT
ws_cover['B27'].hyperlink = 'https://opensheets.co.uk'
ws_cover['B27'].alignment = AL
ws_cover['B27'].fill = PALE_BLUE_FILL
ws_cover.merge_cells('B27:E27')

ws_cover['B28'] = 'Aligned.tax — file your MTD VAT & Income Tax returns from your spreadsheet'
ws_cover['B28'].font = COVER_LINK_FONT
ws_cover['B28'].hyperlink = 'https://aligned.tax'
ws_cover['B28'].alignment = AL
ws_cover['B28'].fill = PALE_BLUE_FILL
ws_cover.merge_cells('B28:E28')

ws_cover['B29'] = 'Free to use • No sign-up required'
ws_cover['B29'].font = make_font(size=10, italic=True, color=MUTED)
ws_cover['B29'].alignment = AC
ws_cover['B29'].fill = PALE_BLUE_FILL
ws_cover.merge_cells('B29:E29')

# Border around promotions box
for row in range(26, 30):
    for col in ['B', 'C', 'D', 'E']:
        cell = ws_cover[f'{col}{row}']
        if row == 26:
            cell.border = Border(top=thin_side(), left=thin_side() if col=='B' else None, right=thin_side() if col=='E' else None, bottom=None)
        elif row == 29:
            cell.border = Border(top=None, left=thin_side() if col=='B' else None, right=thin_side() if col=='E' else None, bottom=thin_side())
        else:
            cell.border = Border(left=thin_side() if col=='B' else None, right=thin_side() if col=='E' else None)

# Row heights for cover
ws_cover.row_dimensions[2].height = 36
ws_cover.row_dimensions[3].height = 18
ws_cover.row_dimensions[5].height = 20
for r in range(6, 13):
    ws_cover.row_dimensions[r].height = 18
ws_cover.row_dimensions[14].height = 20
for r in range(15, 19):
    ws_cover.row_dimensions[r].height = 18
ws_cover.row_dimensions[20].height = 20
ws_cover.row_dimensions[21].height = 22
ws_cover.row_dimensions[22].height = 22
ws_cover.row_dimensions[24].height = 18
ws_cover.row_dimensions[26].height = 24
ws_cover.row_dimensions[27].height = 22
ws_cover.row_dimensions[28].height = 22
ws_cover.row_dimensions[29].height = 18

# ============================================================
# SAVE
# ============================================================
wb.save('Purchase_Order_Template_v4.xlsx')
print('✅ PO v4 built!')
print('   • Delivery Terms dropdown (FOB, Delivered, Ex Works, CIF, DDP, DAP, CFR)')
print('   • Party card: complete border ALL 4 sides on every row')
print('   • Logo: dashed L/R/T + thin B (matching gold standard)')
print('   • Detail values: complete outer box')
print('   • Table body: BLUE input cols (B,C,D,F), WHITE calculated cols (E,G,H)')
print('   • Input cells EMPTY by default')
print('   • Sheet protection enabled')
print('   • Cover sheet included (first sheet)')
