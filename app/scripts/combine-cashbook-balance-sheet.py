#!/usr/bin/env python3
"""Combine Cash Book and Balance Sheet templates into one workbook."""

from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def copy_sheet(source_ws, target_wb, title=None):
    """Copy an openpyxl worksheet into a different workbook."""
    new_title = title or source_ws.title
    # Ensure unique title
    if new_title in target_wb.sheetnames:
        base = new_title
        i = 2
        new_title = f"{base}_{i}"
        while new_title in target_wb.sheetnames:
            i += 1
            new_title = f"{base}_{i}"

    target_ws = target_wb.create_sheet(title=new_title)

    # Copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    # Copy row heights
    for row_num, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = dim.height

    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Copy cell values and styles
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]
            target_cell.value = cell.value
            target_cell.number_format = copy(cell.number_format)
            if cell.font:
                target_cell.font = copy(cell.font)
            if cell.fill:
                target_cell.fill = copy(cell.fill)
            if cell.border:
                target_cell.border = copy(cell.border)
            if cell.alignment:
                target_cell.alignment = copy(cell.alignment)

    # Copy print area and other page setup if present
    if source_ws.print_area:
        target_ws.print_area = source_ws.print_area

    return target_ws


def main():
    cash = load_workbook('public/templates/Cash_Book_Income_Tracker.xlsx')
    bs = load_workbook('public/templates/Balance_Sheet_Template.xlsx')

    combined = Workbook()
    combined.remove(combined.active)

    for sheet in cash.worksheets:
        copy_sheet(sheet, combined)

    for sheet in bs.worksheets:
        # Rename the Balance Sheet Cover to avoid confusion
        title = sheet.title
        if title == 'Cover':
            title = 'BS Cover'
        copy_sheet(sheet, combined, title=title)

    output_path = 'public/templates/Cash_Book_With_Balance_Sheet.xlsx'
    combined.save(output_path)
    print(f'Created {output_path} with sheets: {combined.sheetnames}')


if __name__ == '__main__':
    main()
